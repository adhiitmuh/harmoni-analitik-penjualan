#!/usr/bin/env python3
"""
Harmoni Analitik — HTML Generator
Jalankan: python3 generate_harmoni_app.py
Output:   Harmoni_Analitik_App.html (di folder yang sama)

Mode 1 (Olsera): Taruh file export Olsera (.xlsx per bulan) di folder 'data_olsera/'
Mode 2 (Legacy): Taruh Analitik_v3.xlsx di folder yang sama
"""

import sys
import os
import json
import re
import glob

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# ─── KONFIGURASI ─────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
OLSERA_DIR  = os.path.join(SCRIPT_DIR, "data_olsera")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "Harmoni_Analitik_App.html")

# Fallback ke Analitik_v3.xlsx jika folder data_olsera tidak ada
EXCEL_PATH = os.path.join(SCRIPT_DIR, "Analitik_v3.xlsx")
if not os.path.exists(EXCEL_PATH):
    for c in [os.path.expanduser("~/Analitik_v3.xlsx"), os.path.expanduser("~/Desktop/Analitik_v3.xlsx")]:
        if os.path.exists(c):
            EXCEL_PATH = c
            break

# Mapping Kategori → Divisi (sesuai data Harmoni)
DIVISI_MAP = {
    'AKSESORIS':              'Young Harmonis',
    'CUSTOM':                 'Custom Orders',
    'PERL. PERBAKIN':         'Young Harmonis',
    'PERL. TNI':              'Young Harmonis',
    'PERLENGKAPAN PASKIBRA':  'Young Harmonis',
    'PERLENGKAPAN PELAYARAN': 'Young Harmonis',
    'PERLENGKAPAN PNS':       'Young Harmonis',
    'PERLENGKAPAN POL PP':    'Young Harmonis',
    'PERLENGKAPAN POLRI':     'Young Harmonis',
    'PERLENGKAPAN PRAMUKA':   'Young Harmonis',
    'PERLENGKAPAN SECURITY':  'Young Harmonis',
    'PERLENGKAPAN SEKOLAH':   'Young Harmonis',
    'PERLENGKAPAN TNI':       'Young Harmonis',
    'SEPATU PDH & PDL':       'Young Harmonis',
    'SERAGAM KARNAVAL':       'Young Harmonis',
}

MONTH_ID  = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'Mei',6:'Jun',
             7:'Jul',8:'Agt',9:'Sep',10:'Okt',11:'Nov',12:'Des'}
MONTH_NUM = {v:k for k,v in MONTH_ID.items()}

def bulan_dari_filename(fname):
    m = re.search(r'(\d{4})-(\d{2})-\d{2}', fname)
    if m:
        return f"{MONTH_ID[int(m.group(2))]}-{m.group(1)[2:]}"
    return None

def bulan_sort_key(b):
    parts = b.split('-')
    return (int('20' + parts[1]), MONTH_NUM.get(parts[0], 0))

# ─── BACA DATA ───────────────────────────────────────────────────────────────

from collections import defaultdict

def safe_num(v, default=0):
    if v is None or v == '—' or v == '' or str(v).strip() == '': return default
    try: return float(v)
    except: return default

def safe_str(v):
    if v is None: return ''
    return str(v).strip()

raw_rows = []

olsera_files = sorted(glob.glob(os.path.join(OLSERA_DIR, "*.xlsx"))) if os.path.isdir(OLSERA_DIR) else []

if olsera_files:
    # ── Mode Olsera: baca file export per bulan ───────────────────────────────
    print(f"📂 Mode Olsera — {len(olsera_files)} file ditemukan di {OLSERA_DIR}")
    for fpath in olsera_files:
        fname = os.path.basename(fpath)
        bulan = bulan_dari_filename(fname)
        if not bulan:
            print(f"  ⚠️  Skip {fname} — tidak bisa baca bulan dari nama file")
            continue
        wb_tmp = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws_tmp = wb_tmp.active
        headers = None
        n = 0
        for row in ws_tmp.iter_rows(values_only=True):
            if not row or row[0] is None: continue
            if str(row[0]).lower() == 'product':
                headers = [str(h).lower().strip() if h else '' for h in row]
                continue
            if not headers: continue
            d = dict(zip(headers, row))
            sku = safe_str(d.get('sku', ''))
            if not sku: continue
            qty = int(safe_num(d.get('sold qty', 0)))
            if qty <= 0: continue
            omzet  = safe_num(d.get('total sales amount', 0))
            hpp    = safe_num(d.get('total cost price', 0))
            profit = safe_num(d.get('profit', 0))
            produk = safe_str(d.get('product', ''))
            varian = safe_str(d.get('variant', ''))
            kat    = safe_str(d.get('group', '')).upper()
            divisi = DIVISI_MAP.get(kat, 'Young Harmonis')
            raw_rows.append({
                'bulan': bulan, 'divisi': divisi, 'kategori': kat,
                'produk': produk, 'varian': varian, 'sku': sku,
                'qty': qty, 'omzet': omzet, 'hpp': hpp, 'profit': profit,
            })
            n += 1
        wb_tmp.close()
        print(f"  ✅ {fname} → {bulan} ({n} baris)")

else:
    # ── Mode Legacy: baca dari sheet Data Mentah di Analitik_v3.xlsx ──────────
    print(f"📂 Mode Legacy — membaca: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if 'Data Mentah' in s), None)
    if not sheet_name:
        print("ERROR: Sheet 'Data Mentah' tidak ditemukan dan folder data_olsera/ kosong.")
        sys.exit(1)
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    header_idx = next((i for i, r in enumerate(all_rows) if r[0] == 'Bulan'), None)
    if header_idx is None:
        print("ERROR: Baris header 'Bulan' tidak ditemukan.")
        sys.exit(1)
    for row in all_rows[header_idx + 1:]:
        bulan = safe_str(row[0])
        if not bulan: continue
        sku = safe_str(row[5])
        if not sku:
            # CUSTOM rows tidak punya SKU individual — pakai synthetic key
            if safe_str(row[2]).upper() == 'CUSTOM':
                sku = 'CUSTOM'
            else:
                continue
        qty = int(safe_num(row[6]))
        omzet  = safe_num(row[10])
        hpp    = safe_num(row[11])
        profit = safe_num(row[12])
        raw_rows.append({
            'bulan': bulan, 'divisi': safe_str(row[1]), 'kategori': safe_str(row[2]),
            'produk': safe_str(row[3]), 'varian': safe_str(row[4]), 'sku': sku,
            'qty': qty, 'omzet': omzet, 'hpp': hpp, 'profit': profit,
        })
    wb.close()

if not raw_rows:
    print("ERROR: Tidak ada data yang dibaca.")
    sys.exit(1)

# ── Susun BULAN_ORDER secara dinamis dari data ────────────────────────────────
BULAN_ORDER = sorted(set(r['bulan'] for r in raw_rows), key=bulan_sort_key)
BULAN_IDX   = {b: i for i, b in enumerate(BULAN_ORDER)}
N_BULAN     = len(BULAN_ORDER)

# Filter raw_rows ke bulan yang dikenal
raw_rows = [r for r in raw_rows if r['bulan'] in BULAN_IDX]

# ── Per Kategori ──────────────────────────────────────────────────────────────
kat_monthly_omzet  = defaultdict(lambda: [0.0] * N_BULAN)
kat_monthly_profit = defaultdict(lambda: [0.0] * N_BULAN)
kat_monthly_hpp    = defaultdict(lambda: [0.0] * N_BULAN)
kat_hpp_total      = defaultdict(float)
kat_profit_total   = defaultdict(float)
kat_divisi         = {}

for r in raw_rows:
    kat = r['kategori']
    idx = BULAN_IDX[r['bulan']]
    kat_monthly_omzet[kat][idx]  += r['omzet']
    kat_monthly_profit[kat][idx] += r['profit']
    kat_monthly_hpp[kat][idx]    += r['hpp']
    kat_hpp_total[kat]    += r['hpp']
    kat_profit_total[kat] += r['profit']
    if kat not in kat_divisi:
        kat_divisi[kat] = r['divisi']

grand_total_monthly        = [0.0] * N_BULAN
grand_total_monthly_profit = [0.0] * N_BULAN
grand_total_monthly_hpp    = [0.0] * N_BULAN
for kat in kat_monthly_omzet:
    for i in range(N_BULAN):
        grand_total_monthly[i]        += kat_monthly_omzet[kat][i]
        grand_total_monthly_profit[i] += kat_monthly_profit[kat][i]
        grand_total_monthly_hpp[i]    += kat_monthly_hpp[kat][i]

bulan_labels = BULAN_ORDER[:]
total_omzet_grand = sum(grand_total_monthly)

kategori_data = sorted([{
    'nama': kat,
    'monthly': list(kat_monthly_omzet[kat]),
    'total': sum(kat_monthly_omzet[kat]),
    'avg': sum(kat_monthly_omzet[kat]) / N_BULAN,
    'pct': sum(kat_monthly_omzet[kat]) / total_omzet_grand * 100 if total_omzet_grand else 0,
    'divisi': kat_divisi.get(kat, '—'),
    'monthly_profit': list(kat_monthly_profit[kat]),
    'monthly_hpp':    list(kat_monthly_hpp[kat]),
} for kat in kat_monthly_omzet], key=lambda x: x['total'], reverse=True)

# ── Ringkasan Per Kategori ────────────────────────────────────────────────────
ringkasan = []
for i, k in enumerate(kategori_data):
    kat = k['nama']
    omzet = k['total']
    hpp = kat_hpp_total[kat]
    profit = kat_profit_total[kat]
    margin = profit / omzet * 100 if omzet else 0
    markup = profit / hpp * 100 if hpp else None
    ringkasan.append({
        'no': i + 1, 'nama': kat, 'omzet': omzet,
        'hpp': hpp if hpp else None, 'profit': profit,
        'margin': margin, 'markup': markup, 'divisi': k['divisi'],
    })

grand_hpp = sum(kat_hpp_total.values())
grand_profit = sum(kat_profit_total.values())
grand = {
    'omzet': total_omzet_grand, 'hpp': grand_hpp, 'profit': grand_profit,
    'margin': grand_profit / total_omzet_grand * 100 if total_omzet_grand else 0,
    'markup': grand_profit / grand_hpp * 100 if grand_hpp else 0,
}

# ── Analisis Divisi ───────────────────────────────────────────────────────────
div_monthly = defaultdict(lambda: {'omzet': [0.0]*N_BULAN, 'profit': [0.0]*N_BULAN})
for r in raw_rows:
    d = r['divisi']
    idx = BULAN_IDX[r['bulan']]
    div_monthly[d]['omzet'][idx]  += r['omzet']
    div_monthly[d]['profit'][idx] += r['profit']

yh   = div_monthly.get('Young Harmonis', {'omzet': [0]*N_BULAN, 'profit': [0]*N_BULAN})
cust = div_monthly.get('Custom Orders',  {'omzet': [0]*N_BULAN, 'profit': [0]*N_BULAN})
divisi_data = []
for i, b in enumerate(BULAN_ORDER):
    total_b = yh['omzet'][i] + cust['omzet'][i]
    divisi_data.append({
        'bulan': b,
        'yh_omzet': yh['omzet'][i],
        'yh_profit': yh['profit'][i],
        'yh_margin': yh['profit'][i] / yh['omzet'][i] * 100 if yh['omzet'][i] else 0,
        'cust_omzet': cust['omzet'][i],
        'cust_profit': cust['profit'][i],
        'total': total_b,
        'kontr_yh': yh['omzet'][i] / total_b * 100 if total_b else 0,
    })

# ── Top Produk ────────────────────────────────────────────────────────────────
produk_agg = defaultdict(lambda: {
    'omzet': 0.0, 'qty': 0, 'profit': 0.0, 'hpp': 0.0, 'kategori': '', 'divisi': '',
    'monthly_omzet': [0.0]*N_BULAN, 'monthly_qty': [0]*N_BULAN, 'monthly_profit': [0.0]*N_BULAN,
})
for r in raw_rows:
    key = (r['produk'], r['varian'])
    idx = BULAN_IDX[r['bulan']]
    produk_agg[key]['omzet']  += r['omzet']
    produk_agg[key]['qty']    += r['qty']
    produk_agg[key]['profit'] += r['profit']
    produk_agg[key]['hpp']    += r['hpp']
    produk_agg[key]['monthly_omzet'][idx]  += r['omzet']
    produk_agg[key]['monthly_qty'][idx]    += r['qty']
    produk_agg[key]['monthly_profit'][idx] += r['profit']
    if not produk_agg[key]['kategori']:
        produk_agg[key]['kategori'] = r['kategori']
        produk_agg[key]['divisi']   = r['divisi']

top_list = sorted([{
    'produk': p, 'varian': v,
    'kategori': agg['kategori'], 'omzet': agg['omzet'],
    'qty': agg['qty'], 'profit': agg['profit'],
    'margin': agg['profit'] / agg['omzet'] * 100 if agg['omzet'] else 0,
    'monthly_omzet':  agg['monthly_omzet'],
    'monthly_qty':    agg['monthly_qty'],
    'monthly_profit': agg['monthly_profit'],
} for (p, v), agg in produk_agg.items()], key=lambda x: x['omzet'], reverse=True)

top_produk = [{**p, 'rank': i + 1} for i, p in enumerate(top_list[:50])]

# ── Analisis Masalah (SKU dengan profit negatif per bulan) ────────────────────
sku_bulan_agg = defaultdict(lambda: defaultdict(lambda: {'qty': 0, 'omzet': 0.0, 'hpp': 0.0, 'profit': 0.0}))
sku_meta = {}
for r in raw_rows:
    sku_bulan_agg[r['sku']][r['bulan']]['qty'] += r['qty']
    sku_bulan_agg[r['sku']][r['bulan']]['omzet'] += r['omzet']
    sku_bulan_agg[r['sku']][r['bulan']]['hpp'] += r['hpp']
    sku_bulan_agg[r['sku']][r['bulan']]['profit'] += r['profit']
    if r['sku'] not in sku_meta:
        sku_meta[r['sku']] = {'produk': r['produk'], 'varian': r['varian'], 'kategori': r['kategori'], 'divisi': r['divisi']}

masalah_raw = []
for sku, bulan_dict in sku_bulan_agg.items():
    if sku == 'CUSTOM': continue
    meta = sku_meta[sku]
    for bulan, vals in bulan_dict.items():
        if vals['omzet'] > 0 and vals['profit'] < 0:
            margin = vals['profit'] / vals['omzet'] * 100
            status = 'HPP > Omzet' if vals['hpp'] > vals['omzet'] else 'Profit Negatif'
            rek = 'Cek harga jual / diskon berlebih' if vals['hpp'] > vals['omzet'] else 'Perlu investigasi HPP'
            masalah_raw.append({
                'sku': sku, 'kat': meta['kategori'], 'produk': meta['produk'],
                'varian': meta['varian'], 'bulan': bulan, 'qty': vals['qty'],
                'omzet': vals['omzet'], 'hpp': vals['hpp'], 'profit': vals['profit'],
                'margin': margin, 'status': status, 'rek': rek,
            })

masalah_raw.sort(key=lambda x: x['profit'])
masalah_sku = [{**m, 'no': i + 1} for i, m in enumerate(masalah_raw)]

# ── Stok & Rekomendasi ────────────────────────────────────────────────────────
sku_monthly_qty = defaultdict(lambda: [0] * 14)
sku_info = {}
for r in raw_rows:
    sku = r['sku']
    idx = BULAN_IDX[r['bulan']]
    sku_monthly_qty[sku][idx] += r['qty']
    if sku not in sku_info:
        sku_info[sku] = {
            'produk': r['produk'], 'varian': r['varian'],
            'kategori': r['kategori'], 'divisi': r['divisi'],
            'harga_unit': round(r['omzet'] / r['qty'], 0) if r['qty'] > 0 else 0,
            'hpp_unit': round(r['hpp'] / r['qty'], 0) if r['qty'] > 0 else 0,
        }

stok_rekomendasi = []
for sku, qty_per_bulan in sku_monthly_qty.items():
    if sku == 'CUSTOM': continue
    info = sku_info[sku]
    total_qty = sum(qty_per_bulan)
    months_active = sum(1 for q in qty_per_bulan if q > 0)
    avg_per_bulan = round(total_qty / max(months_active, 1), 1)

    prev3   = sum(qty_per_bulan[max(0, N_BULAN-6):max(0, N_BULAN-3)])
    recent3 = sum(qty_per_bulan[max(0, N_BULAN-3):])
    recent3_avg = round(recent3 / 3, 1)
    trend_pct = (recent3 - prev3) / prev3 * 100 if prev3 > 0 else (100 if recent3 > 0 else 0)

    if trend_pct > 20:
        trend = 'naik'
    elif trend_pct < -20:
        trend = 'turun'
    else:
        trend = 'stabil'

    # ── Deteksi musiman ──────────────────────────────────────────────────────
    # Rata-rata per bulan kalender (1=Jan..12=Des), bisa lintas tahun
    from collections import defaultdict as _dd
    cal_qty = _dd(list)
    for i, b in enumerate(BULAN_ORDER):
        mn = MONTH_NUM.get(b.split('-')[0], 0)
        cal_qty[mn].append(qty_per_bulan[i])
    cal_avg = {m: sum(v)/len(v) for m, v in cal_qty.items() if v}
    peak_month_num = max(cal_avg, key=cal_avg.get) if cal_avg else 0
    peak_avg       = cal_avg.get(peak_month_num, 0)
    overall_avg    = total_qty / max(N_BULAN, 1)
    seasonal_ratio = peak_avg / overall_avg if overall_avg > 0 else 1
    is_seasonal    = seasonal_ratio >= 2.5 and total_qty >= 10 and months_active >= 3
    peak_month_name = MONTH_ID.get(peak_month_num, '')

    if is_seasonal:
        # Rekomendasi = rata-rata bulan puncak + buffer 20%
        rek_qty = max(1, round(peak_avg * 1.2))
    elif trend == 'naik':
        rek_qty = max(1, round(recent3_avg * 1.3))
    elif trend == 'turun':
        rek_qty = max(1, round(recent3_avg * 0.9))
    else:
        rek_qty = max(1, round(recent3_avg * 1.1))

    last_sold = next((b for b in reversed(BULAN_ORDER) if sku_monthly_qty[sku][BULAN_IDX[b]] > 0), '')

    stok_rekomendasi.append({
        'sku': sku, 'produk': info['produk'], 'varian': info['varian'],
        'kategori': info['kategori'], 'divisi': info['divisi'],
        'harga_unit': info['harga_unit'], 'hpp_unit': info['hpp_unit'],
        'total_qty': total_qty, 'months_active': months_active,
        'avg_per_bulan': avg_per_bulan, 'recent3_avg': recent3_avg,
        'trend': trend, 'trend_pct': round(trend_pct, 1),
        'rek_qty': rek_qty, 'last_sold': last_sold,
        'monthly': list(qty_per_bulan),
        'is_seasonal': is_seasonal,
        'peak_month': peak_month_name,
        'peak_avg': round(peak_avg, 1),
    })

stok_rekomendasi.sort(key=lambda x: x['total_qty'], reverse=True)

print(f"✅ Data dibaca: {len(kategori_data)} kategori, {len(top_produk)} top produk, {len(masalah_sku)} SKU masalah, {len(stok_rekomendasi)} SKU stok")

# ─── HITUNG KPI ───────────────────────────────────────────────────────────────
total_omzet = grand['omzet'] if grand else sum(r['omzet'] for r in ringkasan)
total_profit = grand['profit'] if grand else sum(r['profit'] for r in ringkasan)
total_margin = grand['margin'] if grand else (total_profit / total_omzet * 100 if total_omzet else 0)

# Hitung per-tahun secara dinamis
bulan_per_tahun = defaultdict(list)
for i, b in enumerate(BULAN_ORDER):
    tahun = '20' + b.split('-')[1]
    bulan_per_tahun[tahun].append(grand_total_monthly[i])

tahun_list = sorted(bulan_per_tahun.keys())
omzet_per_tahun = {t: sum(bulan_per_tahun[t]) for t in tahun_list}
omzet_2025 = omzet_per_tahun.get('2025', 0)
avg_bulanan = omzet_2025 / max(len(bulan_per_tahun.get('2025', [1])), 1)
ytd_2026 = omzet_per_tahun.get('2026', 0)
n_2026 = len(bulan_per_tahun.get('2026', []))

bulan_terkuat_idx = grand_total_monthly.index(max(grand_total_monthly)) if grand_total_monthly else 0
bulan_terkuat = bulan_labels[bulan_terkuat_idx] if bulan_labels else 'N/A'
omzet_terkuat = max(grand_total_monthly) if grand_total_monthly else 0

yh_total_omzet = sum(d['yh_omzet'] for d in divisi_data)
cust_total_omzet = sum(d['cust_omzet'] for d in divisi_data)
yh_total_profit = sum(d['yh_profit'] for d in divisi_data)
yh_pct = yh_total_omzet / total_omzet * 100 if total_omzet else 0
yh_margin_avg = yh_total_profit / yh_total_omzet * 100 if yh_total_omzet else 0

total_kerugian = sum(m['profit'] for m in masalah_sku)

def fmt_m(v):
    if v >= 1e9: return f"Rp {v/1e9:.2f} M"
    if v >= 1e6: return f"Rp {v/1e6:.1f} Jt"
    return f"Rp {v:,.0f}"

# ─── JSON DATA UNTUK HTML ─────────────────────────────────────────────────────
js_bulan         = json.dumps(bulan_labels)
js_total_bulanan = json.dumps(grand_total_monthly)
js_total_profit_bulanan = json.dumps(grand_total_monthly_profit)
js_total_hpp_bulanan    = json.dumps(grand_total_monthly_hpp)
js_yh_omzet  = json.dumps([d['yh_omzet']  for d in divisi_data])
js_yh_profit = json.dumps([d['yh_profit'] for d in divisi_data])
js_yh_margin = json.dumps([round(d['yh_margin'], 2) for d in divisi_data])
js_cust_omzet = json.dumps([d['cust_omzet'] for d in divisi_data])
js_kontr_yh   = json.dumps([round(d['kontr_yh'], 2) for d in divisi_data])

js_kat_names          = json.dumps([k['nama']           for k in kategori_data])
js_kat_monthly        = json.dumps([k['monthly']        for k in kategori_data])
js_kat_monthly_profit = json.dumps([k['monthly_profit'] for k in kategori_data])
js_kat_monthly_hpp    = json.dumps([k['monthly_hpp']    for k in kategori_data])
js_kat_divisi         = json.dumps([k['divisi']         for k in kategori_data])

js_ringkasan   = json.dumps(ringkasan)
js_divisi_data = json.dumps(divisi_data)
js_top_produk  = json.dumps(top_produk)
js_masalah     = json.dumps(masalah_sku)

js_stok        = json.dumps(stok_rekomendasi)
js_bulan_order = json.dumps(BULAN_ORDER)
js_n_bulan     = N_BULAN
js_divisi_map  = json.dumps(DIVISI_MAP)

from datetime import datetime
generated_at = datetime.now().strftime('%d %b %Y %H:%M')
year_btns = ''.join(f"<button class=\"pbtn\" onclick=\"setPeriod('{y}',this)\">{y}</button>" for y in tahun_list)

# ─── GENERATE HTML ────────────────────────────────────────────────────────────
print("🔨 Membuat HTML...")

html = f"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Harmoni Analitik Stok dan Penjualan</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  :root{{--g:#034543;--g2:#055a57;--g3:#0a7874;--g4:#1a8b88;--beige:#FFFBD5;--beige2:#f5f0b8;--white:#FAFAFA;--black:#282828;--gray:#64748b}}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Inter',sans-serif;background:var(--white);color:var(--black);overflow-x:hidden}}
  .header{{background:linear-gradient(135deg,var(--g) 0%,var(--g2) 50%,var(--g3) 100%);color:white;padding:18px 28px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 4px 20px rgba(3,69,67,.3)}}
  .header-logo{{font-size:1.6rem;font-weight:800;color:white;letter-spacing:-.5px;line-height:1}}
  .header-logo span{{font-size:.8rem;font-weight:400;opacity:.7;margin-left:4px}}
  .header .sub{{font-size:.75rem;opacity:.6;margin-top:3px;font-weight:500}}
  .header .badge{{background:rgba(255,251,213,.18);border:1px solid rgba(255,251,213,.3);padding:4px 14px;border-radius:20px;font-size:.73rem;color:var(--beige)}}
  .nav{{background:white;border-bottom:2px solid rgba(3,69,67,.08);padding:0 24px;display:flex;gap:0;overflow-x:auto}}
  .nav-btn{{padding:13px 19px;font-size:.82rem;font-weight:600;cursor:pointer;border:none;background:transparent;color:#94a3b8;border-bottom:3px solid transparent;white-space:nowrap;transition:all .2s;font-family:'Inter',sans-serif}}
  .nav-btn:hover{{color:var(--g)}}
  .nav-btn.active{{color:var(--g);border-bottom-color:var(--g);font-weight:700}}
  .main{{padding:22px;max-width:1400px;margin:0 auto}}
  .tab-content{{display:none}}.tab-content.active{{display:block}}
  .kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:16px;margin-bottom:22px}}
  .kpi-card{{background:white;border-radius:14px;padding:20px;box-shadow:0 1px 8px rgba(3,69,67,.08);border:1px solid rgba(3,69,67,.08);position:relative;overflow:hidden}}
  .kpi-card::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:linear-gradient(90deg,var(--g),var(--g4))}}
  .kpi-icon{{font-size:1.3rem;margin-bottom:7px}}
  .kpi-label{{font-size:.68rem;color:#94a3b8;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-bottom:5px}}
  .kpi-value{{font-size:1.5rem;font-weight:800;color:var(--g);line-height:1}}
  .kpi-sub{{font-size:.68rem;color:var(--gray);margin-top:5px;font-weight:500}}
  .chart-grid{{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-bottom:22px}}
  .chart-grid-3{{display:grid;grid-template-columns:2fr 1fr;gap:18px;margin-bottom:22px}}
  @media(max-width:900px){{.chart-grid,.chart-grid-3{{grid-template-columns:1fr}}}}
  .card{{background:white;border-radius:14px;padding:20px;box-shadow:0 1px 8px rgba(3,69,67,.08);border:1px solid rgba(3,69,67,.08)}}
  .card-title{{font-size:.88rem;font-weight:700;color:var(--black);margin-bottom:14px;padding-bottom:11px;border-bottom:1px solid rgba(3,69,67,.08);display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}}
  .card-full{{margin-bottom:22px}}
  .table-wrap{{overflow-x:auto}}
  table{{width:100%;border-collapse:collapse;font-size:.8rem}}
  th{{background:linear-gradient(180deg,#f5fffe,#edf9f8);padding:10px 12px;text-align:left;font-weight:700;color:var(--g);font-size:.72rem;text-transform:uppercase;letter-spacing:.4px;border-bottom:2px solid rgba(3,69,67,.12);white-space:nowrap}}
  td{{padding:10px 12px;border-bottom:1px solid rgba(3,69,67,.06);color:var(--black)}}
  tr:hover td{{background:rgba(255,251,213,.35)}}
  .tr{{text-align:right;font-family:'Plus Jakarta Sans','SF Mono',monospace;font-variant-numeric:tabular-nums}}
  .num{{font-family:'Plus Jakarta Sans',sans-serif;font-variant-numeric:tabular-nums;font-weight:600}}
  .tc{{text-align:center}}
  .badge-red{{background:#fee2e2;color:#dc2626;padding:2px 8px;border-radius:5px;font-size:.7rem;font-weight:700}}
  .badge-orange{{background:#ffedd5;color:#c2410c;padding:2px 8px;border-radius:5px;font-size:.7rem;font-weight:700}}
  .badge-yellow{{background:var(--beige);color:#854d0e;padding:2px 8px;border-radius:5px;font-size:.7rem;font-weight:700}}
  .badge-green{{background:rgba(3,69,67,.1);color:var(--g);padding:2px 8px;border-radius:5px;font-size:.7rem;font-weight:700}}
  .badge-blue{{background:rgba(3,69,67,.08);color:var(--g);padding:2px 8px;border-radius:5px;font-size:.7rem;font-weight:700}}
  .mbar-wrap{{display:flex;align-items:center;gap:8px}}
  .mbar-bg{{flex:1;background:rgba(3,69,67,.1);border-radius:4px;height:5px;min-width:50px}}
  .mbar-fill{{height:5px;border-radius:4px}}
  .insight-box{{background:linear-gradient(135deg,rgba(3,69,67,.06),rgba(255,251,213,.4));border-left:4px solid var(--g);border-radius:0 10px 10px 0;padding:12px 16px;margin-bottom:20px;font-size:.82rem;color:var(--g)}}
  .insight-box ul{{padding-left:16px;margin-top:6px}}
  .insight-box li{{margin-bottom:4px}}
  .insight-warn{{background:linear-gradient(135deg,#fff7ed,#ffedd5);border-color:#ea580c;color:#9a3412}}
  .divisi-stat{{display:flex;gap:16px;margin-bottom:20px;flex-wrap:wrap}}
  .div-block{{flex:1;min-width:220px;padding:16px;border-radius:12px}}
  .div-yh{{background:linear-gradient(135deg,rgba(3,69,67,.07),rgba(3,69,67,.03));border:1px solid rgba(3,69,67,.18)}}
  .div-cust{{background:linear-gradient(135deg,var(--beige),var(--beige2));border:1px solid rgba(3,69,67,.15)}}
  .div-label{{font-size:.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-bottom:7px;color:var(--g)}}
  .div-val{{font-size:1.3rem;font-weight:800;color:var(--g)}}
  .div-sub{{font-size:.72rem;color:var(--gray);margin-top:4px;font-weight:500}}
  .filter-row{{display:flex;gap:8px;align-items:center;margin-bottom:14px;flex-wrap:wrap}}
  .filter-row select,.filter-row input{{padding:6px 11px;border:1px solid rgba(3,69,67,.2);border-radius:7px;font-size:.8rem;background:white;color:var(--black);cursor:pointer;font-family:'Inter',sans-serif;outline:none}}
  .filter-row select:focus{{border-color:var(--g)}}
  .bRek{{background:linear-gradient(135deg,var(--beige),var(--beige2));color:var(--g);padding:3px 10px;border-radius:6px;font-weight:800;font-size:.8rem;border:1px solid rgba(3,69,67,.15)}}
  .gen-info{{text-align:right;font-size:.7rem;color:#94a3b8;padding:6px 24px 16px}}
  .sc-hide{{display:none!important}}
  #stokTableWrap.stok-hide-skat .sc-skat,#stokTableWrap.stok-hide-savg .sc-savg,#stokTableWrap.stok-hide-savg3 .sc-savg3,#stokTableWrap.stok-hide-strnd .sc-strnd,#stokTableWrap.stok-hide-strndp .sc-strndp,#stokTableWrap.stok-hide-stipe .sc-stipe,#stokTableWrap.stok-hide-sqty .sc-sqty,#stokTableWrap.stok-hide-srek .sc-srek,#stokTableWrap.stok-hide-stoko .sc-stoko,#stokTableWrap.stok-hide-sgdng .sc-sgdng,#stokTableWrap.stok-hide-stotal .sc-stotal,#stokTableWrap.stok-hide-sbeli .sc-sbeli,#stokTableWrap.stok-hide-slast .sc-slast,#stokTableWrap.stok-hide-sfinal .sc-sfinal,#stokTableWrap.stok-hide-smodal .sc-smodal,#stokTableWrap.stok-hide-sjual .sc-sjual,#stokTableWrap.stok-hide-suntung .sc-suntung{{display:none!important}}
  #stokKolomPanel{{display:none;position:absolute;background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:12px 14px;z-index:200;box-shadow:0 4px 20px rgba(0,0,0,.14);min-width:220px;right:0;top:calc(100% + 4px)}}
  #stokKolomPanel label{{display:flex;align-items:center;gap:7px;font-size:.79rem;padding:2px 0;cursor:pointer;color:#334155;user-select:none}}
  #stokKolomPanel .kp-group{{font-size:.67rem;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.06em;margin:8px 0 2px;padding-top:6px;border-top:1px solid #f1f5f9}}
  #stokKolomPanel .kp-group:first-of-type{{margin-top:0;border-top:none}}
  .period-bar{{background:white;border-bottom:1px solid rgba(3,69,67,.1);padding:8px 24px;display:flex;align-items:center;gap:8px;flex-wrap:wrap;position:sticky;top:52px;z-index:99}}
  .period-bar span{{font-size:.75rem;font-weight:600;color:#94a3b8;margin-right:2px;white-space:nowrap}}
  .pbtn{{padding:4px 13px;border-radius:20px;border:1.5px solid rgba(3,69,67,.25);background:white;font-size:.75rem;font-weight:600;color:var(--g);cursor:pointer;font-family:'Inter',sans-serif;transition:all .15s}}
  .pbtn:hover{{border-color:var(--g);background:rgba(3,69,67,.06)}}
  .pbtn.active{{background:var(--g);color:white;border-color:var(--g)}}
  .period-badge{{font-size:.7rem;background:rgba(3,69,67,.1);color:var(--g);padding:2px 10px;border-radius:20px;font-weight:700;white-space:nowrap}}
</style>
</head>
<body>

<div class="header">
  <div>
    <div class="header-logo">harmoni <span>· Analitik Stok dan Penjualan</span></div>
    <div class="sub">{bulan_labels[0] if bulan_labels else 'N/A'} – {bulan_labels[-1] if bulan_labels else 'N/A'} &nbsp;·&nbsp; {len(bulan_labels)} Bulan Data Aktual</div>
  </div>
  <div style="display:flex;flex-direction:column;align-items:flex-end;gap:6px">
    <div class="badge">Update: {generated_at}</div>
  </div>
</div>

<div class="nav">
  <button class="nav-btn active" onclick="showTab('dashboard',this)">📊 Dashboard</button>
  <button class="nav-btn" onclick="showTab('kategori',this)">📂 Per Kategori</button>
  <button class="nav-btn" onclick="showTab('divisi',this)">🏢 Analisis Divisi</button>
  <button class="nav-btn" onclick="showTab('produk',this)">🏆 Top Produk</button>
  <button class="nav-btn" onclick="showTab('masalah',this)">🚨 Analisis Masalah</button>
  <button class="nav-btn" onclick="showTab('inventori',this)">🏪 Inventori Stok</button>
  <button class="nav-btn" onclick="showTab('stok',this)">📦 Rekomendasi</button>
</div>

<div class="period-bar">
  <span>Periode:</span>
  <button class="pbtn active" onclick="setPeriod('all',this)">Semua</button>
  {year_btns}
  <button class="pbtn" onclick="setPeriod('12m',this)">12 Bln</button>
  <button class="pbtn" onclick="setPeriod('6m',this)">6 Bln</button>
  <button class="pbtn" onclick="setPeriod('3m',this)">3 Bln</button>
  <select id="bulanPicker" onchange="setPeriodBulan(this)" style="padding:4px 10px;border-radius:20px;border:1.5px solid rgba(3,69,67,.25);font-size:.75rem;font-weight:600;color:var(--g);background:white;cursor:pointer;font-family:'Inter',sans-serif;outline:none">
    <option value="" selected>-- Pilih Bulan --</option>
    {''.join(f'<option value="{b}">{b}</option>' for b in bulan_labels)}
  </select>
  <span id="periodLabel" class="period-badge">Semua {len(bulan_labels)} Bulan</span>
</div>

<div class="main">

<!-- DASHBOARD -->
<div id="tab-dashboard" class="tab-content active">
  <div class="kpi-grid">
    <div class="kpi-card"><div class="kpi-icon">💰</div><div class="kpi-label" id="kpi-omzet-label">Total Omzet ({len(bulan_labels)} Bln)</div><div class="kpi-value" id="kpi-omzet">{fmt_m(total_omzet)}</div><div class="kpi-sub" id="kpi-omzet-sub">{bulan_labels[0] if bulan_labels else ''} – {bulan_labels[-1] if bulan_labels else ''}</div></div>
    <div class="kpi-card"><div class="kpi-icon">📈</div><div class="kpi-label">Total Profit</div><div class="kpi-value" id="kpi-profit">{fmt_m(total_profit)}</div><div class="kpi-sub" id="kpi-margin-sub">Margin rata-rata: {total_margin:.1f}%</div></div>
    <div class="kpi-card"><div class="kpi-icon">📅</div><div class="kpi-label" id="kpi-avg-label">Rata-rata/Bulan</div><div class="kpi-value" id="kpi-avg">{fmt_m(avg_bulanan)}</div><div class="kpi-sub" id="kpi-avg-sub">dari {len(bulan_labels)} bulan</div></div>
    <div class="kpi-card"><div class="kpi-icon">🏆</div><div class="kpi-label">Bulan Terkuat</div><div class="kpi-value" id="kpi-best-month">{bulan_terkuat}</div><div class="kpi-sub" id="kpi-best-omzet">{fmt_m(omzet_terkuat)} omzet</div></div>
    <div class="kpi-card"><div class="kpi-icon">📊</div><div class="kpi-label" id="kpi-yh-label">Young Harmonis</div><div class="kpi-value" id="kpi-yh">{fmt_m(yh_total_omzet)}</div><div class="kpi-sub" id="kpi-yh-sub">{yh_pct:.1f}% dari total</div></div>
    <div class="kpi-card"><div class="kpi-icon">⚠️</div><div class="kpi-label">SKU Margin Negatif</div><div class="kpi-value">{len(masalah_sku)} SKU</div><div class="kpi-sub">Total kerugian: {fmt_m(abs(total_kerugian))}</div></div>
  </div>
  <div class="chart-grid-3">
    <div class="card"><div class="card-title">📈 Tren Omzet Bulanan</div><div style="position:relative;height:280px"><canvas id="cTren"></canvas></div></div>
    <div class="card"><div class="card-title">🍩 Distribusi Omzet Per Kategori</div><div style="position:relative;height:280px"><canvas id="cDonut"></canvas></div></div>
  </div>
  <div class="chart-grid">
    <div class="card"><div class="card-title">💼 Young Harmonis vs Custom Orders</div><div style="position:relative;height:240px"><canvas id="cDivisiBar"></canvas></div></div>
    <div class="card"><div class="card-title">📊 Omzet Per Kategori Utama</div><div style="position:relative;height:240px"><canvas id="cKatBar"></canvas></div></div>
  </div>
  <div class="insight-box">
    <strong>📌 Insight Kunci</strong>
    <ul>
      <li>{bulan_terkuat} adalah bulan tertinggi — {fmt_m(omzet_terkuat)} omzet (puncak musim sekolah/PASKIBRA).</li>
      <li>Young Harmonis menyumbang <strong>{yh_pct:.1f}%</strong> omzet total, Custom Orders {100-yh_pct:.1f}%.</li>
      <li>Custom Orders memiliki margin <strong>100%</strong> (tidak ada HPP tercatat).</li>
      <li>{len(masalah_sku)} SKU dijual di bawah HPP — total kerugian {fmt_m(abs(total_kerugian))}.</li>
    </ul>
  </div>
</div>

<!-- PER KATEGORI -->
<div id="tab-kategori" class="tab-content">
  <div class="card card-full">
    <div class="card-title">📊 Tren Omzet Per Kategori</div>
    <div class="filter-row">
      <label style="font-size:.82rem;font-weight:500;color:#475569">Kategori:</label>
      <select id="katFilter" onchange="updateKatChart()">
        {''.join(f'<option value="{i}">{k["nama"]}</option>' for i, k in enumerate(kategori_data[:8]))}
        <option value="99">SEMUA (Total)</option>
      </select>
    </div>
    <div style="position:relative;height:260px"><canvas id="cKatLine"></canvas></div>
  </div>
  <div class="card card-full">
    <div class="card-title">📋 Ringkasan Total Per Kategori</div>
    <div class="table-wrap"><table>
      <thead><tr><th>No</th><th>Kategori</th><th>Divisi</th><th class="tr">Total Omzet</th><th class="tr">Total HPP</th><th class="tr">Total Profit</th><th>Margin</th><th class="tr">Markup</th></tr></thead>
      <tbody id="tbKat"></tbody>
      <tfoot><tr style="font-weight:700;background:#f8fafc">
        <td colspan="3" style="padding:10px 12px">GRAND TOTAL</td>
        <td class="tr" style="padding:10px 12px">Rp {grand['omzet']:,.0f}</td>
        <td class="tr" style="padding:10px 12px">Rp {grand.get('hpp',0):,.0f}</td>
        <td class="tr" style="padding:10px 12px">Rp {grand['profit']:,.0f}</td>
        <td style="padding:10px 12px"><span class="badge-green">{grand['margin']:.1f}%</span></td>
        <td class="tr" style="padding:10px 12px">{grand['markup']:.1f}%</td>
      </tr></tfoot>
    </table></div>
  </div>
</div>

<!-- ANALISIS DIVISI -->
<div id="tab-divisi" class="tab-content">
  <div class="divisi-stat">
    <div class="div-block div-yh">
      <div class="div-label">🔵 Young Harmonis</div>
      <div class="div-val">{fmt_m(yh_total_omzet)}</div>
      <div class="div-sub">{yh_pct:.1f}% dari total omzet · Profit {fmt_m(yh_total_profit)} · Margin {yh_margin_avg:.1f}%</div>
    </div>
    <div class="div-block div-cust">
      <div class="div-label">🟢 Custom Orders</div>
      <div class="div-val">{fmt_m(cust_total_omzet)}</div>
      <div class="div-sub">{100-yh_pct:.1f}% dari total omzet · Profit {fmt_m(cust_total_omzet)} · Margin 100%</div>
    </div>
  </div>
  <div class="card card-full"><div class="card-title">📈 Tren Bulanan — YH vs Custom</div><div style="position:relative;height:280px"><canvas id="cDivisiLine"></canvas></div></div>
  <div class="card card-full">
    <div class="card-title">📋 Detail Per Bulan</div>
    <div class="table-wrap"><table>
      <thead><tr><th>Bulan</th><th class="tr">YH Omzet</th><th class="tr">YH Profit</th><th>YH Margin</th><th class="tr">Custom Omzet</th><th class="tr">Custom Profit</th><th class="tr">Kontr. YH</th><th class="tr">Kontr. Custom</th></tr></thead>
      <tbody id="tbDiv"></tbody>
    </table></div>
  </div>
</div>

<!-- TOP PRODUK -->
<div id="tab-produk" class="tab-content">
  <div class="chart-grid">
    <div class="card"><div class="card-title">🏆 Top 10 — Omzet Tertinggi</div><div style="position:relative;height:320px"><canvas id="cTopOmzet"></canvas></div></div>
    <div class="card"><div class="card-title">📊 Top 10 — Margin Tertinggi</div><div style="position:relative;height:320px"><canvas id="cTopMargin"></canvas></div></div>
  </div>
  <div class="card card-full">
    <div class="card-title" style="display:flex;justify-content:space-between;align-items:center">
      <span>📋 Top {len(top_produk)} Produk Terlaris</span>
      <div class="filter-row" style="margin:0;gap:8px">
        <select id="topKatFilter" onchange="filterTopProduk()">
          <option value="">Semua Kategori</option>
          {''.join(f'<option>{k["nama"]}</option>' for k in ringkasan)}
        </select>
        <select id="topSortFilter" onchange="filterTopProduk()">
          <option value="omzet">Urutkan: Omzet</option>
          <option value="qty">Urutkan: Qty Terjual</option>
          <option value="profit">Urutkan: Profit</option>
          <option value="margin">Urutkan: Margin</option>
        </select>
      </div>
    </div>
    <div class="table-wrap"><table>
      <thead><tr><th>No</th><th>Produk</th><th>Varian</th><th>Kategori</th><th class="tr">Total Omzet</th><th class="tr">Total Qty</th><th class="tr">Total Profit</th><th>Margin</th></tr></thead>
      <tbody id="tbTop"></tbody>
    </table></div>
  </div>
</div>

<!-- ANALISIS MASALAH -->
<div id="tab-masalah" class="tab-content">
  <div class="insight-box" style="background:#fff7ed;border-color:#ea580c;color:#9a3412">
    <strong>⚠️ Ringkasan Masalah</strong>
    <ul>
      <li>{len(masalah_sku)} SKU dengan profit <strong>negatif</strong> — dijual di bawah HPP. Total kerugian: <strong>{fmt_m(abs(total_kerugian))}</strong></li>
      <li>Perlu evaluasi harga jual, pengecekan HPP, atau pertimbangan discontinue produk.</li>
    </ul>
  </div>
  <div class="card card-full">
    <div class="card-title" style="display:flex;justify-content:space-between;align-items:center">
      <span>🔴 SKU dengan Profit Negatif ({len(masalah_sku)} SKU)</span>
      <div class="filter-row" style="margin:0">
        <select id="masFilter" onchange="filterMasalah()">
          <option value="">Semua Kategori</option>
          {''.join(f'<option>{k}</option>' for k in sorted(set(m["kat"] for m in masalah_sku)))}
        </select>
      </div>
    </div>
    <div class="table-wrap"><table>
      <thead><tr><th>No</th><th>SKU</th><th>Kategori</th><th>Produk</th><th>Varian</th><th>Bulan</th><th class="tr">Qty</th><th class="tr">Omzet</th><th class="tr">HPP</th><th class="tr">Profit</th><th class="tr">Margin</th><th>Status</th><th>Rekomendasi</th></tr></thead>
      <tbody id="tbMas"></tbody>
    </table></div>
  </div>
</div>

<!-- STOK & REKOMENDASI -->
<div id="tab-inventori" class="tab-content">
  <div class="card card-full">
    <div class="card-title" style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px">
      <span>🏪 Inventori Stok — Toko &amp; Gudang</span>
      <div style="display:flex;flex-wrap:wrap;gap:8px;align-items:center">
        <input type="file" id="invTokoInput" accept=".xlsx" multiple style="display:none" onchange="uploadStokFisik(this.files,'toko')">
        <input type="file" id="invGudangInput" accept=".xlsx" multiple style="display:none" onchange="uploadStokFisik(this.files,'gudang')">
        <button onclick="document.getElementById('invTokoInput').click()" style="padding:6px 14px;background:#0369a1;color:#fff;border:none;border-radius:8px;font-size:.8rem;font-weight:600;cursor:pointer;font-family:'Inter',sans-serif">🏪 Upload Stok Toko</button>
        <span id="invTokoLabel" style="font-size:.75rem;color:#0369a1;font-weight:600"></span>
        <button onclick="document.getElementById('invGudangInput').click()" style="padding:6px 14px;background:#7c3aed;color:#fff;border:none;border-radius:8px;font-size:.8rem;font-weight:600;cursor:pointer;font-family:'Inter',sans-serif">🏭 Upload Stok Gudang</button>
        <span id="invGudangLabel" style="font-size:.75rem;color:#7c3aed;font-weight:600"></span>
      </div>
    </div>
    <div class="filter-row" style="margin:10px 0 8px;gap:8px">
      <input id="invSearch" type="text" placeholder="Cari produk / SKU..." oninput="filterInventori()" style="padding:7px 12px;border:1px solid #e2e8f0;border-radius:8px;font-size:.82rem;width:200px">
      <select id="invKatFilter" onchange="filterInventori()" style="padding:7px 10px;border:1px solid #e2e8f0;border-radius:8px;font-size:.82rem">
        <option value="">Semua Kategori</option>
        {''.join(f'<option value="{k}">{k}</option>' for k in sorted(set(r['kategori'] for r in stok_rekomendasi)))}
      </select>
      <select id="invSrcFilter" onchange="filterInventori()" style="padding:7px 10px;border:1px solid #e2e8f0;border-radius:8px;font-size:.82rem">
        <option value="">Toko &amp; Gudang</option>
        <option value="toko">Hanya Toko</option>
        <option value="gudang">Hanya Gudang</option>
      </select>
    </div>
    <div id="invCount" style="font-size:.8rem;color:#64748b;margin-bottom:8px"></div>
    <div id="invEmpty" style="text-align:center;padding:40px;color:#94a3b8;display:none">
      <div style="font-size:2rem;margin-bottom:8px">📦</div>
      <div style="font-weight:600">Belum ada data stok</div>
      <div style="font-size:.82rem;margin-top:4px">Upload file Olsera toko dan/atau gudang dulu</div>
    </div>
    <div class="table-wrap" id="invTableWrap" style="display:none;max-height:70vh;overflow:auto">
    <table style="border-collapse:separate;border-spacing:0">
      <thead style="position:-webkit-sticky;position:sticky;top:0;z-index:4"><tr>
        <th style="position:sticky;left:0;z-index:3;background:#034543;min-width:38px">No</th>
        <th style="position:sticky;left:38px;z-index:3;background:#034543;min-width:90px">SKU</th>
        <th style="position:sticky;left:128px;z-index:3;background:#034543;min-width:160px">Produk</th>
        <th style="position:sticky;left:288px;z-index:3;background:#034543;min-width:100px;border-right:2px solid #0a7874">Varian</th>
        <th>Kategori</th>
        <th class="tr" style="background:#e0f2fe;color:#0369a1">Qty Toko</th>
        <th class="tr" style="background:#e0f2fe;color:#0369a1;min-width:110px">Modal Toko</th>
        <th class="tr" style="background:#f3e8ff;color:#7c3aed">Qty Gudang</th>
        <th class="tr" style="background:#f3e8ff;color:#7c3aed;min-width:110px">Modal Gudang</th>
        <th class="tr" style="background:#f0fdf4;color:#166534;font-weight:700">Total Stok</th>
        <th class="tr" style="background:#fef9c3;color:#854d0e;min-width:110px">Total Modal</th>
        <th class="tr" style="background:#dcfce7;color:#166534;min-width:110px">Total Nilai Jual</th>
        <th class="tr" style="background:#eff6ff;color:#1d4ed8;min-width:110px">Pot. Untung</th>
        <th class="tc" style="min-width:80px">Exclude</th>
      </tr></thead>
      <tbody id="tbInv"></tbody>
    </table></div>
    <div id="invFooter" style="display:none;margin-top:12px;padding:14px 18px;background:#f8fafc;border-radius:10px;border:1px solid #e2e8f0">
      <div style="display:flex;flex-wrap:wrap;gap:24px;justify-content:space-between;align-items:center">
        <div style="display:flex;gap:24px;flex-wrap:wrap">
          <div>
            <div style="font-size:.72rem;color:#0369a1;font-weight:700;text-transform:uppercase;letter-spacing:.04em;margin-bottom:2px">Qty Toko</div>
            <div id="inv_qty_toko" style="font-family:'Plus Jakarta Sans',sans-serif;font-variant-numeric:tabular-nums;font-weight:800;font-size:1rem;color:#0369a1">—</div>
          </div>
          <div>
            <div style="font-size:.72rem;color:#7c3aed;font-weight:700;text-transform:uppercase;letter-spacing:.04em;margin-bottom:2px">Qty Gudang</div>
            <div id="inv_qty_gudang" style="font-family:'Plus Jakarta Sans',sans-serif;font-variant-numeric:tabular-nums;font-weight:800;font-size:1rem;color:#7c3aed">—</div>
          </div>
          <div>
            <div style="font-size:.72rem;color:#166534;font-weight:700;text-transform:uppercase;letter-spacing:.04em;margin-bottom:2px">Total Stok</div>
            <div id="inv_qty_total" style="font-family:'Plus Jakarta Sans',sans-serif;font-variant-numeric:tabular-nums;font-weight:800;font-size:1rem;color:#166534">—</div>
          </div>
        </div>
        <div style="display:flex;gap:24px;flex-wrap:wrap">
          <div style="text-align:right">
            <div style="font-size:.72rem;color:#854d0e;font-weight:700;text-transform:uppercase;letter-spacing:.04em;margin-bottom:2px">Total Modal</div>
            <div id="inv_modal_total" style="font-family:'Plus Jakarta Sans',sans-serif;font-variant-numeric:tabular-nums;font-weight:800;font-size:1rem;color:#854d0e">—</div>
          </div>
          <div style="text-align:right">
            <div style="font-size:.72rem;color:#166534;font-weight:700;text-transform:uppercase;letter-spacing:.04em;margin-bottom:2px">Total Nilai Jual</div>
            <div id="inv_jual_total" style="font-family:'Plus Jakarta Sans',sans-serif;font-variant-numeric:tabular-nums;font-weight:800;font-size:1rem;color:#166534">—</div>
          </div>
          <div style="text-align:right;border-left:2px solid #e2e8f0;padding-left:18px">
            <div style="font-size:.72rem;color:#1d4ed8;font-weight:700;text-transform:uppercase;letter-spacing:.04em;margin-bottom:2px">Pot. Untung</div>
            <div id="inv_untung_total" style="font-family:'Plus Jakarta Sans',sans-serif;font-variant-numeric:tabular-nums;font-weight:800;font-size:1rem;color:#1d4ed8">—</div>
          </div>
        </div>
      </div>
    </div>
    <div id="invKatSummary" style="display:none;margin-top:16px">
      <div style="font-size:.85rem;font-weight:700;color:#034543;margin-bottom:8px">📂 Ringkasan Per Kategori</div>
      <div class="table-wrap"><table>
        <thead><tr>
          <th>Kategori</th>
          <th class="tr">Jumlah SKU</th>
          <th class="tr">Total Stok</th>
          <th class="tr" style="background:#fef9c3;color:#854d0e">Total Modal</th>
          <th class="tr" style="background:#dcfce7;color:#166534">Total Nilai Jual</th>
          <th class="tr" style="background:#eff6ff;color:#1d4ed8">Pot. Untung</th>
          <th class="tr" style="background:#fef3c7;color:#92400e">% dari Total Modal</th>
        </tr></thead>
        <tbody id="tbInvKat"></tbody>
      </table></div>
    </div>
  </div>
</div>

<div id="tab-stok" class="tab-content">
  <div class="insight-box">
    <strong>📦 Cara Baca Rekomendasi Stok</strong>
    <ul>
      <li><strong>🗓 Musiman</strong> = penjualan terkonsentrasi di bulan tertentu (puncak ≥2.5× rata-rata) → rekomendasi pakai rata-rata bulan puncak + 20% buffer.</li>
      <li><strong>Naik ↑</strong> = tumbuh &gt;20% vs 3 bulan sebelumnya → sediakan 30% lebih banyak dari rata-rata recent.</li>
      <li><strong>Stabil →</strong> = perubahan ±20% → buffer 10% dari rata-rata recent.</li>
      <li><strong>Turun ↓</strong> = turun &gt;20% → kurangi 10% dari rata-rata recent.</li>
      <li>Filter per kategori, tren, atau cari produk di kotak pencarian.</li>
    </ul>
  </div>
  <div class="card card-full">
    <div class="card-title" style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px">
      <span>📦 Rekomendasi Stok Per SKU</span>
      <div style="display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-top:6px">
        <label style="padding:6px 14px;background:#034543;color:#fff;border:none;border-radius:8px;font-size:.8rem;font-weight:600;cursor:pointer;font-family:'Inter',sans-serif">
          📤 Upload Data Penjualan
          <input type="file" id="olseraInput" accept=".xlsx" multiple style="display:none" onchange="handleOlseraUpload(this.files)">
        </label>
        <span id="uploadStatus" style="font-size:.75rem;color:#64748b;font-weight:500"></span>
        <span id="stokTokoLabel" style="font-size:.75rem;color:#0369a1;font-weight:600;padding:4px 10px;background:#eff6ff;border-radius:8px"></span>
        <span id="stokGudangLabel" style="font-size:.75rem;color:#7c3aed;font-weight:600;padding:4px 10px;background:#f3e8ff;border-radius:8px"></span>
        <button onclick="downloadStokCSV()" style="padding:6px 14px;background:#475569;color:#fff;border:none;border-radius:8px;font-size:.8rem;font-weight:600;cursor:pointer;font-family:'Inter',sans-serif">⬇ CSV</button>
        <button onclick="downloadStokXLSX()" style="padding:6px 14px;background:#166534;color:#fff;border:none;border-radius:8px;font-size:.8rem;font-weight:600;cursor:pointer;font-family:'Inter',sans-serif">⬇ Excel</button>
      </div>
      <div class="filter-row" style="margin:6px 0 0;gap:8px;position:relative">
        <input id="stokSearch" type="text" placeholder="Cari produk..." oninput="filterStok()" style="padding:7px 12px;border:1px solid #e2e8f0;border-radius:8px;font-size:.82rem;width:200px">
        <select id="stokBeliFilter" onchange="filterStok()">
          <option value="">Semua</option>
          <option value="beli">Perlu Beli</option>
          <option value="cukup">Stok Cukup</option>
        </select>
        <select id="stokKatFilter" onchange="filterStok()">
          <option value="">Semua Kategori</option>
          {''.join(f'<option>{k["nama"]}</option>' for k in ringkasan if k["nama"] != "CUSTOM")}
        </select>
        <select id="stokTrendFilter" onchange="filterStok()">
          <option value="">Semua Tren</option>
          <option value="naik">📈 Tren Naik</option>
          <option value="stabil">➡️ Tren Stabil</option>
          <option value="turun">📉 Tren Turun</option>
          <option value="musiman">🗓 Musiman</option>
        </select>
        <button id="btnStokKolom" onclick="toggleStokKolomPanel(event)" style="padding:6px 12px;background:#f1f5f9;color:#475569;border:1px solid #e2e8f0;border-radius:8px;font-size:.8rem;font-weight:600;cursor:pointer;font-family:'Inter',sans-serif;white-space:nowrap">⚙ Kolom</button>
        <div id="stokKolomPanel">
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
            <span style="font-size:.8rem;font-weight:700;color:#1e293b">Tampilkan Kolom</span>
            <div style="display:flex;gap:5px">
              <button onclick="stokKolomAll(true)" style="font-size:.7rem;padding:2px 8px;border:1px solid #e2e8f0;border-radius:5px;cursor:pointer;background:#f8fafc;color:#475569">Semua</button>
              <button onclick="stokKolomAll(false)" style="font-size:.7rem;padding:2px 8px;border:1px solid #e2e8f0;border-radius:5px;cursor:pointer;background:#f8fafc;color:#475569">Sembunyikan</button>
            </div>
          </div>
          <div class="kp-group">Analisis</div>
          <label><input type="checkbox" data-col="skat" checked onchange="toggleStokCol('skat',this.checked)"> Kategori</label>
          <label><input type="checkbox" data-col="savg" checked onchange="toggleStokCol('savg',this.checked)"> Avg/Bln</label>
          <label><input type="checkbox" data-col="savg3" checked onchange="toggleStokCol('savg3',this.checked)"> Avg 3Bln Terakhir</label>
          <label><input type="checkbox" data-col="strnd" checked onchange="toggleStokCol('strnd',this.checked)"> Tren</label>
          <label><input type="checkbox" data-col="strndp" checked onchange="toggleStokCol('strndp',this.checked)"> Tren %</label>
          <label><input type="checkbox" data-col="stipe" checked onchange="toggleStokCol('stipe',this.checked)"> Tipe</label>
          <label><input type="checkbox" data-col="sqty" checked onchange="toggleStokCol('sqty',this.checked)"> Total Qty Terjual</label>
          <div class="kp-group">Stok</div>
          <label><input type="checkbox" data-col="srek" checked onchange="toggleStokCol('srek',this.checked)"> Rek. Stok</label>
          <label><input type="checkbox" data-col="stoko" checked onchange="toggleStokCol('stoko',this.checked)"> Stok Toko</label>
          <label><input type="checkbox" data-col="sgdng" checked onchange="toggleStokCol('sgdng',this.checked)"> Stok Gudang</label>
          <label><input type="checkbox" data-col="stotal" checked onchange="toggleStokCol('stotal',this.checked)"> Total Stok</label>
          <label><input type="checkbox" data-col="sbeli" checked onchange="toggleStokCol('sbeli',this.checked)"> Perlu Beli</label>
          <label><input type="checkbox" data-col="slast" checked onchange="toggleStokCol('slast',this.checked)"> Terakhir Terjual</label>
          <div class="kp-group">Finansial</div>
          <label><input type="checkbox" data-col="sfinal" checked onchange="toggleStokCol('sfinal',this.checked)"> Final Order</label>
          <label><input type="checkbox" data-col="smodal" checked onchange="toggleStokCol('smodal',this.checked)"> Est. Modal</label>
          <label><input type="checkbox" data-col="sjual" checked onchange="toggleStokCol('sjual',this.checked)"> Est. Nilai Jual</label>
          <label><input type="checkbox" data-col="suntung" checked onchange="toggleStokCol('suntung',this.checked)"> Est. Untung</label>
        </div>
      </div>
    </div>
    <div id="stokCount" style="font-size:.78rem;color:#64748b;margin-bottom:10px"></div>
    <div class="table-wrap" id="stokTableWrap" style="max-height:70vh;overflow:auto"><table style="border-collapse:separate;border-spacing:0">
      <thead style="position:-webkit-sticky;position:sticky;top:0;z-index:4"><tr>
        <th style="position:sticky;left:0;z-index:3;background:#034543;min-width:38px">No</th>
        <th style="position:sticky;left:38px;z-index:3;background:#034543;min-width:90px">SKU</th>
        <th style="position:sticky;left:128px;z-index:3;background:#034543;min-width:160px">Produk</th>
        <th style="position:sticky;left:288px;z-index:3;background:#034543;min-width:100px;border-right:2px solid #0a7874">Varian</th>
        <th class="sc-skat">Kategori</th>
        <th class="tr sc-savg">Avg/Bln</th>
        <th class="tr sc-savg3">Avg 3Bln Terakhir</th><th class="sc-strnd">Tren</th><th class="sc-strndp">Tren %</th>
        <th class="sc-stipe">Tipe</th>
        <th class="tr sc-sqty">Total Qty Terjual</th>
        <th class="tr sc-srek" style="background:#fef9c3;color:#854d0e">Rek. Stok</th>
        <th class="tr sc-stoko" style="background:#e0f2fe;color:#0369a1">Stok Toko</th>
        <th class="tr sc-sgdng" style="background:#f3e8ff;color:#7c3aed">Stok Gudang</th>
        <th class="tr sc-stotal" style="background:#f0fdf4;color:#166534">Total Stok</th>
        <th class="tr sc-sbeli" style="background:#fef2f2;color:#dc2626">Perlu Beli</th>
        <th class="sc-slast">Terakhir Terjual</th>
        <th class="sc-sfinal" style="background:#dcfce7;color:#166534;min-width:90px">Final Order</th>
        <th class="tr sc-smodal" style="background:#fef9c3;color:#854d0e;min-width:110px">Est. Modal</th>
        <th class="tr sc-sjual" style="background:#dcfce7;color:#166534;min-width:110px">Est. Nilai Jual</th>
        <th class="tr sc-suntung" style="background:#eff6ff;color:#1d4ed8;min-width:110px">Est. Untung</th>
      </tr></thead>
      <tbody id="tbStok"></tbody>
    </table></div>
    <div style="margin-top:12px;padding:14px 18px;background:#fef9c3;border-radius:10px;border:1px solid #fde68a">
      <div style="display:flex;flex-wrap:wrap;gap:24px;justify-content:flex-end;align-items:center">
        <div style="text-align:right">
          <div style="font-size:.72rem;color:#92400e;font-weight:700;text-transform:uppercase;letter-spacing:.04em;margin-bottom:2px">Total Modal</div>
          <div id="totalBudgetVal" style="font-family:'Plus Jakarta Sans',sans-serif;font-variant-numeric:tabular-nums;font-weight:800;font-size:1rem;color:#854d0e">—</div>
        </div>
        <div style="text-align:right">
          <div style="font-size:.72rem;color:#166534;font-weight:700;text-transform:uppercase;letter-spacing:.04em;margin-bottom:2px">Total Nilai Jual</div>
          <div id="totalJualVal" style="font-family:'Plus Jakarta Sans',sans-serif;font-variant-numeric:tabular-nums;font-weight:800;font-size:1rem;color:#166534">—</div>
        </div>
        <div style="text-align:right">
          <div style="font-size:.72rem;color:#1d4ed8;font-weight:700;text-transform:uppercase;letter-spacing:.04em;margin-bottom:2px">Est. Profit</div>
          <div id="totalProfitVal" style="font-family:'Plus Jakarta Sans',sans-serif;font-variant-numeric:tabular-nums;font-weight:800;font-size:1rem;color:#1d4ed8">—</div>
        </div>
        <div style="text-align:right;border-left:2px solid #fde068;padding-left:18px">
          <div style="font-size:.72rem;color:#7c3aed;font-weight:700;text-transform:uppercase;letter-spacing:.04em;margin-bottom:2px">Est. Margin</div>
          <div id="totalMarginVal" style="font-family:'Plus Jakarta Sans',sans-serif;font-variant-numeric:tabular-nums;font-weight:800;font-size:1rem;color:#7c3aed">—</div>
        </div>
      </div>
    </div>
  </div>
</div>

</div>
<div class="gen-info">Generated by generate_harmoni_app.py · {generated_at}</div>

<script>
let BULAN = {js_bulan};
let TOTAL_BULANAN         = {js_total_bulanan};
let TOTAL_PROFIT_BULANAN  = {js_total_profit_bulanan};
let TOTAL_HPP_BULANAN     = {js_total_hpp_bulanan};
let YH_OMZET   = {js_yh_omzet};
let YH_PROFIT  = {js_yh_profit};
let YH_MARGIN  = {js_yh_margin};
let CUST_OMZET = {js_cust_omzet};
let KONTR_YH   = {js_kontr_yh};
let KAT_NAMES          = {js_kat_names};
let KAT_MONTHLY        = {js_kat_monthly};
let KAT_MONTHLY_PROFIT = {js_kat_monthly_profit};
let KAT_MONTHLY_HPP    = {js_kat_monthly_hpp};
let KAT_DIVISI         = {js_kat_divisi};
let RINGKASAN   = {js_ringkasan};
let DIVISI_DATA = {js_divisi_data};
let TOP_PRODUK  = {js_top_produk};
let MASALAH_SKU = {js_masalah};
let STOK_DATA   = {js_stok};
const BULAN_ORDER = {js_bulan_order};
const DIVISI_MAP  = {js_divisi_map};

// ── Olsera Upload ─────────────────────────────────────────────────────────────
const MONTH_ID_JS = {{1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'Mei',6:'Jun',7:'Jul',8:'Agt',9:'Sep',10:'Okt',11:'Nov',12:'Des'}};
const MONTH_NUM_JS = {{'Jan':1,'Feb':2,'Mar':3,'Apr':4,'Mei':5,'Jun':6,'Jul':7,'Agt':8,'Sep':9,'Okt':10,'Nov':11,'Des':12}};

function bulanDariFile(fname){{
  const m = fname.match(/(\d{{4}})-(\d{{2}})-\d{{2}}/);
  if(!m) return null;
  return MONTH_ID_JS[parseInt(m[2])]+'-'+m[1].slice(2);
}}

function bulanSortKey(b){{
  const p=b.split('-');
  return parseInt('20'+p[1])*100+(MONTH_NUM_JS[p[0]]||0);
}}

function handleOlseraUpload(files){{
  if(!files||!files.length) return;
  const st=document.getElementById('uploadStatus');
  st.textContent='⏳ Memuat...';
  if(typeof XLSX==='undefined'){{
    const s=document.createElement('script');
    s.src='https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js';
    s.onload=()=>processOlseraFiles(files,st);
    s.onerror=()=>{{st.textContent='⚠️ Gagal memuat library Excel. Cek koneksi internet.';}};
    document.head.appendChild(s);
    return;
  }}
  processOlseraFiles(files,st);
}}

function processOlseraFiles(files,st){{
  st.textContent='⏳ Membaca '+files.length+' file...';
  const rawRows=[];
  let done=0;
  Array.from(files).forEach(file=>{{
    const bulan=bulanDariFile(file.name);
    if(!bulan){{
      st.textContent='⚠️ Nama file tidak dikenali: '+file.name;
      done++; if(done===files.length) finishUpload(rawRows,files.length,st);
      return;
    }}
    const reader=new FileReader();
    reader.onload=e=>{{
      try{{
        const wb=XLSX.read(e.target.result,{{type:'array'}});
        const ws=wb.Sheets[wb.SheetNames[0]];
        const rows=XLSX.utils.sheet_to_json(ws,{{header:1}});
        const hRow=rows.findIndex(r=>r[0]&&String(r[0]).toLowerCase().trim()==='product');
        if(hRow<0){{ done++; if(done===files.length) finishUpload(rawRows,files.length,st); return; }}
        const hdrs=rows[hRow].map(h=>h?String(h).toLowerCase().trim():'');
        const ci=n=>hdrs.indexOf(n);
        const iP=ci('product'),iV=ci('variant'),iG=ci('group'),iS=ci('sku');
        const iQ=ci('sold qty'),iO=ci('total sales amount'),iH=ci('total cost price'),iPr=ci('profit');
        for(let r=hRow+1;r<rows.length;r++){{
          const row=rows[r];
          const sku=String(row[iS]||'').trim();
          if(!sku) continue;
          const qty=parseFloat(row[iQ])||0;
          if(qty<=0) continue;
          const omzet=parseFloat(row[iO])||0;
          const hpp=parseFloat(row[iH])||0;
          const profit=parseFloat(row[iPr])||0;
          const kat=String(row[iG]||'').toUpperCase().trim();
          const divisi=DIVISI_MAP[kat]||'Young Harmonis';
          rawRows.push({{bulan,divisi,kategori:kat,produk:String(row[iP]||'').trim(),
            varian:String(row[iV]||'').trim(),sku,qty,omzet,hpp,profit}});
        }}
      }}catch(err){{console.error(err);}}
      done++;
      if(done===files.length) finishUpload(rawRows,files.length,st);
    }};
    reader.readAsArrayBuffer(file);
  }});
}}

function finishUpload(rawRows,nFiles,st){{
  if(!rawRows.length){{ st.textContent='⚠️ Data tidak terbaca'; return; }}
  rebuildAllData(rawRows);
  const nBulan=[...new Set(rawRows.map(r=>r.bulan))].length;
  st.textContent='✅ '+nFiles+' file · '+nBulan+' bulan · '+rawRows.length+' transaksi';
  // reset period filter ke All
  document.querySelectorAll('.pbtn').forEach(b=>b.classList.remove('active'));
  document.querySelector('.pbtn')?.classList.add('active');
  document.getElementById('periodLabel').textContent='Semua '+BULAN.length+' Bulan';
}}

function rebuildAllData(rawRows){{
  const allBulan=([...new Set(rawRows.map(r=>r.bulan))]).sort((a,b)=>bulanSortKey(a)-bulanSortKey(b));
  const bidx=Object.fromEntries(allBulan.map((b,i)=>[b,i]));
  const N=allBulan.length;

  // Per-kategori
  const katOmzet={{}}, katProfit={{}}, katHpp={{}}, katDiv={{}};
  const divOmzet={{}}, divProfit={{}};
  const prodAgg={{}};
  const skuMonthly={{}}, skuMeta={{}};
  const skuBulanAgg={{}};

  rawRows.forEach(r=>{{
    const i=bidx[r.bulan];
    // kategori
    if(!katOmzet[r.kategori]){{ katOmzet[r.kategori]=new Array(N).fill(0); katProfit[r.kategori]=new Array(N).fill(0); katHpp[r.kategori]=new Array(N).fill(0); katDiv[r.kategori]=r.divisi; }}
    katOmzet[r.kategori][i]+=r.omzet; katProfit[r.kategori][i]+=r.profit; katHpp[r.kategori][i]+=r.hpp;
    // divisi
    if(!divOmzet[r.divisi]){{ divOmzet[r.divisi]=new Array(N).fill(0); divProfit[r.divisi]=new Array(N).fill(0); }}
    divOmzet[r.divisi][i]+=r.omzet; divProfit[r.divisi][i]+=r.profit;
    // produk
    const pk=r.produk+'||'+r.varian;
    if(!prodAgg[pk]){{ prodAgg[pk]={{produk:r.produk,varian:r.varian,kategori:r.kategori,monthly_omzet:new Array(N).fill(0),monthly_qty:new Array(N).fill(0),monthly_profit:new Array(N).fill(0)}}; }}
    prodAgg[pk].monthly_omzet[i]+=r.omzet; prodAgg[pk].monthly_qty[i]+=r.qty; prodAgg[pk].monthly_profit[i]+=r.profit;
    // sku
    if(!skuMonthly[r.sku]){{ skuMonthly[r.sku]=new Array(N).fill(0); skuMeta[r.sku]={{produk:r.produk,varian:r.varian,kategori:r.kategori,divisi:r.divisi,hpp_unit:r.qty>0?Math.round(r.hpp/r.qty):0,harga_unit:r.qty>0?Math.round(r.omzet/r.qty):0}}; }}
    skuMonthly[r.sku][i]+=r.qty;
    // masalah
    const mk=r.sku+'||'+r.bulan;
    if(!skuBulanAgg[mk]){{ skuBulanAgg[mk]={{sku:r.sku,kat:r.kategori,produk:r.produk,varian:r.varian,bulan:r.bulan,qty:0,omzet:0,hpp:0,profit:0}}; }}
    skuBulanAgg[mk].qty+=r.qty; skuBulanAgg[mk].omzet+=r.omzet; skuBulanAgg[mk].hpp+=r.hpp; skuBulanAgg[mk].profit+=r.profit;
  }});

  const totalBulan=new Array(N).fill(0), totalProfit=new Array(N).fill(0), totalHpp=new Array(N).fill(0);
  Object.values(katOmzet).forEach(a=>a.forEach((v,i)=>totalBulan[i]+=v));
  Object.values(katProfit).forEach(a=>a.forEach((v,i)=>totalProfit[i]+=v));
  Object.values(katHpp).forEach(a=>a.forEach((v,i)=>totalHpp[i]+=v));

  const yhD=divOmzet['Young Harmonis']||new Array(N).fill(0);
  const yhP=divProfit['Young Harmonis']||new Array(N).fill(0);
  const cuD=divOmzet['Custom Orders']||new Array(N).fill(0);
  const cuP=divProfit['Custom Orders']||new Array(N).fill(0);

  // Build KAT_* arrays sorted by total omzet desc
  const katList=Object.keys(katOmzet).sort((a,b)=>katOmzet[b].reduce((s,v)=>s+v,0)-katOmzet[a].reduce((s,v)=>s+v,0));
  KAT_NAMES=[...katList];
  KAT_MONTHLY=katList.map(k=>[...katOmzet[k]]);
  KAT_MONTHLY_PROFIT=katList.map(k=>[...katProfit[k]]);
  KAT_MONTHLY_HPP=katList.map(k=>[...katHpp[k]]);
  KAT_DIVISI=katList.map(k=>katDiv[k]);
  BULAN=[...allBulan];
  TOTAL_BULANAN=[...totalBulan];
  TOTAL_PROFIT_BULANAN=[...totalProfit];
  TOTAL_HPP_BULANAN=[...totalHpp];
  YH_OMZET=[...yhD]; YH_PROFIT=[...yhP];
  CUST_OMZET=[...cuD];
  YH_MARGIN=allBulan.map((_,i)=>yhD[i]>0?+(yhP[i]/yhD[i]*100).toFixed(2):0);
  KONTR_YH=allBulan.map((_,i)=>{{const t=yhD[i]+cuD[i]; return t>0?+(yhD[i]/t*100).toFixed(2):0;}});
  DIVISI_DATA=allBulan.map((b,i)=>{{const t=yhD[i]+cuD[i]; return {{bulan:b,yh_omzet:yhD[i],yh_profit:yhP[i],yh_margin:yhD[i]>0?yhP[i]/yhD[i]*100:0,cust_omzet:cuD[i],cust_profit:cuP[i],total:t,kontr_yh:t>0?yhD[i]/t*100:0}};}});
  RINGKASAN=katList.map((k,i)=>{{const o=katOmzet[k].reduce((s,v)=>s+v,0),p=katProfit[k].reduce((s,v)=>s+v,0),h=katHpp[k].reduce((s,v)=>s+v,0); return {{no:i+1,nama:k,omzet:o,hpp:h||null,profit:p,margin:o>0?p/o*100:0,markup:h>0?p/h*100:null,divisi:katDiv[k]}};}});
  TOP_PRODUK=Object.values(prodAgg).map(p=>{{const o=p.monthly_omzet.reduce((s,v)=>s+v,0),q=p.monthly_qty.reduce((s,v)=>s+v,0),pr=p.monthly_profit.reduce((s,v)=>s+v,0); return {{...p,omzet:o,qty:q,profit:pr,margin:o>0?pr/o*100:0}};}}).filter(p=>p.omzet>0).sort((a,b)=>b.omzet-a.omzet).map((p,i)=>{{p.rank=i+1;return p;}});
  MASALAH_SKU=Object.values(skuBulanAgg).filter(m=>m.omzet>0&&m.profit<0).sort((a,b)=>a.profit-b.profit).map((m,i)=>{{const mg=m.profit/m.omzet*100; return {{...m,no:i+1,margin:mg,status:m.hpp>m.omzet?'HPP > Omzet':'Profit Negatif',rek:m.hpp>m.omzet?'Cek harga jual / diskon':'Perlu investigasi HPP'}};}});

  // Stok rekomendasi
  const stok=[];
  Object.keys(skuMonthly).forEach(sku=>{{
    const q=skuMonthly[sku], inf=skuMeta[sku];
    const tot=q.reduce((s,v)=>s+v,0), act=q.filter(v=>v>0).length;
    const prev3=q.slice(Math.max(0,N-6),Math.max(0,N-3)).reduce((s,v)=>s+v,0);
    const rec3=q.slice(Math.max(0,N-3)).reduce((s,v)=>s+v,0);
    const r3avg=+(rec3/3).toFixed(1);
    const tpct=prev3>0?(rec3-prev3)/prev3*100:rec3>0?100:0;
    const trend=tpct>20?'naik':tpct<-20?'turun':'stabil';
    const rek=Math.max(1,Math.round(r3avg*(trend==='naik'?1.3:trend==='turun'?0.9:1.1)));
    const last=[...allBulan].reverse().find(b=>skuMonthly[sku][bidx[b]]>0)||'';
    stok.push({{sku,produk:inf.produk,varian:inf.varian,kategori:inf.kategori,divisi:inf.divisi,harga_unit:inf.harga_unit,hpp_unit:inf.hpp_unit,total_qty:tot,months_active:act,avg_per_bulan:+(tot/Math.max(act,1)).toFixed(1),recent3_avg:r3avg,trend,trend_pct:+tpct.toFixed(1),rek_qty:rek,last_sold:last,monthly:[...q]}});
  }});
  STOK_DATA=stok.sort((a,b)=>b.total_qty-a.total_qty);

  // Reset period filter dan re-render
  FILTERED_IDX=BULAN.map((_,i)=>i);
  // Rebuild year buttons
  const yrs=[...new Set(BULAN.map(b=>'20'+b.split('-')[1]))].sort();
  const pbar=document.querySelector('.period-bar');
  const fixBtns=pbar.querySelectorAll('.pbtn'); // All, year btns, 12M, 6M, 3M
  // Remove year buttons (between first and last 3)
  const allPbtns=[...pbar.querySelectorAll('.pbtn')];
  allPbtns.slice(1,allPbtns.length-3).forEach(b=>b.remove());
  const ref=allPbtns[allPbtns.length-3]; // 12M button
  yrs.forEach(y=>{{const b=document.createElement('button');b.className='pbtn';b.textContent=y;b.onclick=()=>setPeriod(y,b);pbar.insertBefore(b,ref);}});
  allPbtns[0].classList.add('active');
  allPbtns.slice(1).forEach(b=>b.classList.remove('active'));
  refreshBulanPicker();

  applyPeriodFilter(FILTERED_IDX);
  renderStok(STOK_DATA);
  updateTotalBudget();
  renderMas(MASALAH_SKU);
}}

// ── Period filter state ───────────────────────────────────────────────────────
let FILTERED_IDX = BULAN.map((_,i)=>i);  // default: all months

function getFilteredIdx(mode){{
  const n = BULAN.length;
  if(mode==='all') return BULAN.map((_,i)=>i);
  if(mode==='3m')  return [...Array(Math.min(3,n)).keys()].map(i=>n-Math.min(3,n)+i);
  if(mode==='6m')  return [...Array(Math.min(6,n)).keys()].map(i=>n-Math.min(6,n)+i);
  if(mode==='12m') return [...Array(Math.min(12,n)).keys()].map(i=>n-Math.min(12,n)+i);
  // year mode e.g. '2025'
  return BULAN.map((b,i)=>b.endsWith('-'+mode.slice(2))?i:-1).filter(i=>i>=0);
}}

function fmtMjs(v){{
  if(!v||v===0)return '—';
  if(v>=1e9)return 'Rp '+(v/1e9).toFixed(2)+' M';
  if(v>=1e6)return 'Rp '+(v/1e6).toFixed(1)+' Jt';
  return 'Rp '+Math.round(v).toLocaleString('id-ID');
}}

function applyPeriodFilter(idx){{
  FILTERED_IDX = idx;
  const fBulan = idx.map(i=>BULAN[i]);
  const fN = fBulan.length;

  // KPI recompute
  const totalOmzet  = idx.reduce((s,i)=>s+TOTAL_BULANAN[i],0);
  const totalProfit = idx.reduce((s,i)=>s+TOTAL_PROFIT_BULANAN[i],0);
  const totalHpp    = idx.reduce((s,i)=>s+TOTAL_HPP_BULANAN[i],0);
  const margin = totalOmzet>0?totalProfit/totalOmzet*100:0;
  const avg = fN>0?totalOmzet/fN:0;
  const yhOmzet   = idx.reduce((s,i)=>s+YH_OMZET[i],0);
  const custOmzet = idx.reduce((s,i)=>s+CUST_OMZET[i],0);
  const yhPct = totalOmzet>0?yhOmzet/totalOmzet*100:0;
  const bestIdx = idx.reduce((bi,i)=>TOTAL_BULANAN[i]>TOTAL_BULANAN[bi]?i:bi, idx[0]??0);

  document.getElementById('kpi-omzet').textContent  = fmtMjs(totalOmzet);
  document.getElementById('kpi-profit').textContent = fmtMjs(totalProfit);
  document.getElementById('kpi-margin-sub').textContent = 'Margin rata-rata: '+margin.toFixed(1)+'%';
  document.getElementById('kpi-avg').textContent    = fmtMjs(avg);
  document.getElementById('kpi-avg-sub').textContent = 'dari '+fN+' bulan terpilih';
  document.getElementById('kpi-best-month').textContent = BULAN[bestIdx]??'—';
  document.getElementById('kpi-best-omzet').textContent  = fmtMjs(TOTAL_BULANAN[bestIdx]??0)+' omzet';
  document.getElementById('kpi-yh').textContent     = fmtMjs(yhOmzet);
  document.getElementById('kpi-yh-sub').textContent = yhPct.toFixed(1)+'% dari total';
  document.getElementById('kpi-omzet-label').textContent = 'Total Omzet ('+fN+' Bln)';
  document.getElementById('kpi-omzet-sub').textContent  = (fBulan[0]??'')+(fBulan.length>1?' – '+(fBulan[fBulan.length-1]??''):'');

  // Update semua chart bertanda 'f' (filtered)
  const fTotalBulan = idx.map(i=>TOTAL_BULANAN[i]);
  const fYH   = idx.map(i=>YH_OMZET[i]);
  const fCust = idx.map(i=>CUST_OMZET[i]);

  [cTren,cDivisiLine].forEach(c=>{{ c.data.labels=fBulan; c.update(); }});
  cTren.data.datasets[0].data = fTotalBulan; cTren.update();
  cDivisiLine.data.datasets[0].data = fYH;
  cDivisiLine.data.datasets[1].data = fCust;
  cDivisiLine.data.labels = fBulan; cDivisiLine.update();

  cDivisiBar.data.labels = fBulan;
  cDivisiBar.data.datasets[0].data = fYH;
  cDivisiBar.data.datasets[1].data = fCust;
  cDivisiBar.update();

  // Donut + KatBar: recompute per-kategori totals for filtered period
  const fKatOmzet = KAT_MONTHLY.map(m=>idx.reduce((s,i)=>s+m[i],0));
  const fKatProfit = KAT_MONTHLY_PROFIT.map(m=>idx.reduce((s,i)=>s+m[i],0));
  const top6 = KAT_NAMES.map((nm,i)=>{{return{{nm,omzet:fKatOmzet[i],profit:fKatProfit[i]}};}})
    .sort((a,b)=>b.omzet-a.omzet).slice(0,6);
  cDonut.data.labels = top6.map(k=>k.nm.replace('PERLENGKAPAN ','').replace('SERAGAM ',''));
  cDonut.data.datasets[0].data = top6.map(k=>k.omzet);
  cDonut.update();
  cKatBar.data.labels = top6.map(k=>k.nm.length>18?k.nm.slice(0,18)+'…':k.nm);
  cKatBar.data.datasets[0].data = top6.map(k=>k.omzet);
  cKatBar.update();

  // KatLine: update with filtered months
  updateKatChart();

  // Top Produk table + charts
  renderTopProduk(idx);

  // Ringkasan table
  renderRingkasan(fKatOmzet, fKatProfit, KAT_MONTHLY_HPP.map(m=>idx.reduce((s,i)=>s+m[i],0)));
}}

function setPeriodBulan(sel){{
  if(!sel.value)return;
  document.querySelectorAll('.pbtn').forEach(b=>b.classList.remove('active'));
  const idx=[BULAN.indexOf(sel.value)].filter(i=>i>=0);
  document.getElementById('periodLabel').textContent=sel.value;
  applyPeriodFilter(idx);
}}

function refreshBulanPicker(){{
  const sel=document.getElementById('bulanPicker');
  sel.innerHTML='<option value="" selected>-- Pilih Bulan --</option>'+BULAN.map(b=>`<option value="${{b}}">${{b}}</option>`).join('');
}}

function setPeriod(mode, btn){{
  document.querySelectorAll('.pbtn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  const idx = getFilteredIdx(mode);
  const label = mode==='all'?'Semua '+BULAN.length+' Bulan':
    mode==='3m'?'3 Bulan Terakhir':mode==='6m'?'6 Bulan Terakhir':
    mode==='12m'?'12 Bulan Terakhir':mode+' ('+idx.length+' Bln)';
  document.getElementById('periodLabel').textContent = label;
  document.getElementById('bulanPicker').value = '';
  applyPeriodFilter(idx);
}}

const COLORS = ['#034543','#1a8b88','#2eb8b4','#c2a000','#ea580c','#7c3aed','#dc2626','#d97706'];

Chart.defaults.font.family="'Inter',-apple-system,BlinkMacSystemFont,sans-serif";
Chart.defaults.font.size=11;

function fmtRp(v){{
  if(v==null||v===0)return '—';
  return 'Rp '+Math.round(v).toLocaleString('id-ID');
}}
function fmtM(v){{
  if(v>=1e9)return 'Rp '+(v/1e9).toFixed(2)+' M';
  if(v>=1e6)return 'Rp '+(v/1e6).toFixed(1)+' Jt';
  return 'Rp '+Math.round(v).toLocaleString('id-ID');
}}
function pct(v,d=1){{return v.toFixed(d)+'%';}}
function mBadge(m){{
  if(m>=60)return`<span class="badge-green">${{pct(m)}}</span>`;
  if(m>=40)return`<span class="badge-blue">${{pct(m)}}</span>`;
  if(m>=25)return`<span class="badge-yellow">${{pct(m)}}</span>`;
  return`<span class="badge-red">${{pct(m)}}</span>`;
}}
function mBar(m,max=100){{
  const p=Math.min(Math.max(m,0),max)/max*100;
  const c=m<25?'#ef4444':m<40?'#f59e0b':m<60?'#034543':'#22c55e';
  return`<div class="mbar-wrap">${{mBadge(m)}}<div class="mbar-bg"><div class="mbar-fill" style="width:${{p}}%;background:${{c}}"></div></div></div>`;
}}

// ── Inventori Stok ────────────────────────────────────────────────────────────
const _infoMap={{}};
STOK_DATA.forEach(s=>{{ _infoMap[s.sku]=s; }});

function isInvExcluded(sku){{ return localStorage.getItem('inv_excl_'+sku)==='1'; }}
function toggleInvExclude(sku){{
  const was=isInvExcluded(sku);
  localStorage.setItem('inv_excl_'+sku, was?'0':'1');
  filterInventori();
}}

function renderInventori(items){{
  const tbody=document.getElementById('tbInv');
  const wrap=document.getElementById('invTableWrap');
  const empty=document.getElementById('invEmpty');
  const footer=document.getElementById('invFooter');
  const katDiv=document.getElementById('invKatSummary');
  const cnt=document.getElementById('invCount');
  if(!items||!items.length){{
    wrap.style.display='none'; empty.style.display='block';
    footer.style.display='none'; katDiv.style.display='none';
    cnt.textContent='';
    return;
  }}
  wrap.style.display='block'; empty.style.display='none';
  footer.style.display='block'; katDiv.style.display='block';
  cnt.textContent=items.length+' produk ditampilkan';
  adjustTableWrapHeight();
  tbody.innerHTML='';
  items.forEach((item,i)=>{{
    const excl=isInvExcluded(item.sku);
    const rowBg=i%2===0?'#fff':'#f8fafc';
    const sid=item.sku.replace(/[^a-zA-Z0-9]/g,'_');
    tbody.innerHTML+=`<tr style="opacity:${{excl?'0.35':'1'}}">
      <td class="tc" style="position:sticky;left:0;z-index:2;background:${{rowBg}};min-width:38px">${{i+1}}</td>
      <td style="position:sticky;left:38px;z-index:2;background:${{rowBg}};min-width:90px"><code style="font-size:.75rem;background:#f1f5f9;padding:1px 4px;border-radius:3px">${{item.sku}}</code></td>
      <td style="position:sticky;left:128px;z-index:2;background:${{rowBg}};min-width:160px;font-weight:600">${{item.produk}}</td>
      <td style="position:sticky;left:288px;z-index:2;background:${{rowBg}};min-width:100px;border-right:2px solid #e2e8f0">${{item.varian}}</td>
      <td><span class="badge-blue" style="font-size:.68rem">${{item.kategori.replace('PERLENGKAPAN ','')}}</span></td>
      <td class="tr num" style="color:#0369a1">${{item.qToko>0?item.qToko.toLocaleString('id-ID'):'—'}}</td>
      <td class="tr num" style="color:#0369a1">${{item.modalToko>0?fmtRp(item.modalToko):(item.qToko>0?'HPP?':'—')}}</td>
      <td class="tr num" style="color:#7c3aed">${{item.qGudang>0?item.qGudang.toLocaleString('id-ID'):'—'}}</td>
      <td class="tr num" style="color:#7c3aed">${{item.modalGudang>0?fmtRp(item.modalGudang):(item.qGudang>0?'HPP?':'—')}}</td>
      <td class="tr num" style="font-weight:700">${{item.qTotal.toLocaleString('id-ID')}}</td>
      <td class="tr num" style="color:#854d0e;font-weight:700">${{item.modalTotal>0?fmtRp(item.modalTotal):'—'}}</td>
      <td class="tr num" style="color:#166534;font-weight:700">${{item.jualTotal>0?fmtRp(item.jualTotal):'—'}}</td>
      <td class="tr num" style="color:#1d4ed8;font-weight:700">${{item.untung>0?fmtRp(item.untung):'—'}}</td>
      <td class="tc"><button onclick="toggleInvExclude('${{item.sku}}')"
        style="padding:2px 10px;border-radius:12px;font-size:.72rem;font-weight:700;cursor:pointer;border:none;
        background:${{excl?'#fee2e2':'#f0fdf4'}};color:${{excl?'#dc2626':'#166534'}}">${{excl?'Excluded':'Sertakan'}}</button></td>
    </tr>`;
  }});
  updateInventoriSummary(items);
}}

function filterInventori(){{
  const allSkus=[...new Set([...Object.keys(STOK_TOKO),...Object.keys(STOK_GUDANG)])];
  if(!allSkus.length){{ renderInventori([]); return; }}
  const q=document.getElementById('invSearch').value.toLowerCase();
  const kat=document.getElementById('invKatFilter').value;
  const src=document.getElementById('invSrcFilter').value;
  const items=allSkus.map(sku=>{{
    const info=_infoMap[sku]||{{produk:sku,varian:'',kategori:'Lainnya',hpp_unit:0,harga_unit:0}};
    const qToko=STOK_TOKO[sku]||0, qGudang=STOK_GUDANG[sku]||0;
    const qTotal=qToko+qGudang;
    const hpp=info.hpp_unit||0, harga=info.harga_unit||0;
    return {{sku,produk:info.produk,varian:info.varian||'',kategori:info.kategori,
      qToko,qGudang,qTotal,hpp,harga,
      modalToko:qToko*hpp,modalGudang:qGudang*hpp,
      modalTotal:qTotal*hpp,jualTotal:qTotal*harga,untung:qTotal*harga-qTotal*hpp}};
  }}).filter(item=>{{
    if(q&&!item.produk.toLowerCase().includes(q)&&!item.sku.toLowerCase().includes(q)&&!item.varian.toLowerCase().includes(q)) return false;
    if(kat&&item.kategori!==kat) return false;
    if(src==='toko'&&item.qToko===0) return false;
    if(src==='gudang'&&item.qGudang===0) return false;
    return true;
  }}).sort((a,b)=>b.qTotal-a.qTotal);
  renderInventori(items);
}}

function updateInventoriSummary(items){{
  const active=items.filter(item=>!isInvExcluded(item.sku));
  const totQToko=active.reduce((s,i)=>s+i.qToko,0);
  const totQGudang=active.reduce((s,i)=>s+i.qGudang,0);
  const totQTotal=active.reduce((s,i)=>s+i.qTotal,0);
  const totModal=active.reduce((s,i)=>s+i.modalTotal,0);
  const totJual=active.reduce((s,i)=>s+i.jualTotal,0);
  const totUntung=totJual-totModal;
  const f=id=>document.getElementById(id);
  if(f('inv_qty_toko')) f('inv_qty_toko').textContent=totQToko.toLocaleString('id-ID');
  if(f('inv_qty_gudang')) f('inv_qty_gudang').textContent=totQGudang.toLocaleString('id-ID');
  if(f('inv_qty_total')) f('inv_qty_total').textContent=totQTotal.toLocaleString('id-ID');
  if(f('inv_modal_total')) f('inv_modal_total').textContent=totModal>0?fmtRp(totModal):'—';
  if(f('inv_jual_total')) f('inv_jual_total').textContent=totJual>0?fmtRp(totJual):'—';
  if(f('inv_untung_total')) f('inv_untung_total').textContent=totUntung>0?fmtRp(totUntung):'—';
  // Per-kategori
  const katMap={{}};
  active.forEach(item=>{{
    const k=item.kategori;
    if(!katMap[k]) katMap[k]={{count:0,qty:0,modal:0,jual:0}};
    katMap[k].count++; katMap[k].qty+=item.qTotal;
    katMap[k].modal+=item.modalTotal; katMap[k].jual+=item.jualTotal;
  }});
  const tb=document.getElementById('tbInvKat');
  if(!tb) return;
  const kats=Object.entries(katMap).sort((a,b)=>b[1].modal-a[1].modal);
  tb.innerHTML=kats.map(([kat,d],i)=>{{
    const pct=totModal>0?(d.modal/totModal*100):0;
    const rowBg=i%2===0?'#fff':'#f8fafc';
    return `<tr style="background:${{rowBg}}">
      <td><span class="badge-blue" style="font-size:.72rem">${{kat.replace('PERLENGKAPAN ','')}}</span></td>
      <td class="tr num">${{d.count}}</td>
      <td class="tr num">${{d.qty.toLocaleString('id-ID')}}</td>
      <td class="tr num" style="color:#854d0e;font-weight:600">${{d.modal>0?fmtRp(d.modal):'—'}}</td>
      <td class="tr num" style="color:#166534;font-weight:600">${{d.jual>0?fmtRp(d.jual):'—'}}</td>
      <td class="tr num" style="color:#1d4ed8;font-weight:600">${{(d.jual-d.modal)>0?fmtRp(d.jual-d.modal):'—'}}</td>
      <td class="tr num" style="color:#92400e">${{pct.toFixed(1)}}%</td>
    </tr>`;
  }}).join('');
}}

function adjustTableWrapHeight(){{
  ['invTableWrap','stokTableWrap'].forEach(function(wid){{
    var wrap=document.getElementById(wid);
    if(!wrap||wrap.style.display==='none')return;
    var rect=wrap.getBoundingClientRect();
    if(rect.top>50)wrap.style.maxHeight=Math.max(200,window.innerHeight-rect.top-16)+'px';
  }});
}}
window.addEventListener('resize',adjustTableWrapHeight);

function showTab(id,btn){{
  document.querySelectorAll('.tab-content').forEach(e=>e.classList.remove('active'));
  document.querySelectorAll('.nav-btn').forEach(e=>e.classList.remove('active'));
  document.getElementById('tab-'+id).classList.add('active');
  btn.classList.add('active');
  window.scrollTo(0,0);
  if(id==='inventori') filterInventori();
  if(id==='inventori'||id==='stok') setTimeout(adjustTableWrapHeight,0);
}}

// Charts (named for period filter updates)
const cTren = new Chart(document.getElementById('cTren'),{{type:'line',data:{{labels:BULAN,datasets:[{{label:'Total Omzet',data:TOTAL_BULANAN,borderColor:'#034543',backgroundColor:'rgba(3,69,67,.08)',fill:true,tension:.4,pointRadius:4}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>' '+fmtM(c.raw)}}}}}},scales:{{y:{{ticks:{{callback:v=>fmtM(v)}},grid:{{color:'#f1f5f9'}}}},x:{{grid:{{display:false}}}}}}}}}});
const cDonut = new Chart(document.getElementById('cDonut'),{{type:'doughnut',data:{{labels:RINGKASAN.slice(0,6).map(r=>r.nama.replace('PERLENGKAPAN ','').replace('SERAGAM ','')),datasets:[{{data:RINGKASAN.slice(0,6).map(r=>r.omzet),backgroundColor:COLORS,borderWidth:2,borderColor:'#fff'}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'right',labels:{{font:{{size:10}},padding:10}}}},tooltip:{{callbacks:{{label:c=>' '+fmtM(c.raw)}}}}}},cutout:'60%'}}}});
const cDivisiBar = new Chart(document.getElementById('cDivisiBar'),{{type:'bar',data:{{labels:BULAN,datasets:[{{label:'Young Harmonis',data:YH_OMZET,backgroundColor:'rgba(3,69,67,.85)',borderRadius:4}},{{label:'Custom Orders',data:CUST_OMZET,backgroundColor:'rgba(194,160,0,.8)',borderRadius:4}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{tooltip:{{callbacks:{{label:c=>' '+fmtM(c.raw)}}}}}},scales:{{x:{{stacked:true,grid:{{display:false}}}},y:{{stacked:true,ticks:{{callback:v=>fmtM(v)}},grid:{{color:'#f1f5f9'}}}}}}}}}});
const cKatBar = new Chart(document.getElementById('cKatBar'),{{type:'bar',data:{{labels:RINGKASAN.slice(0,6).map(r=>r.nama.length>18?r.nama.slice(0,18)+'…':r.nama),datasets:[{{label:'Omzet',data:RINGKASAN.slice(0,6).map(r=>r.omzet),backgroundColor:COLORS,borderRadius:6}}]}},options:{{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>' '+fmtM(c.raw)}}}}}},scales:{{x:{{ticks:{{callback:v=>fmtM(v)}},grid:{{color:'#f1f5f9'}}}},y:{{grid:{{display:false}}}}}}}}}});
const cDivisiLine = new Chart(document.getElementById('cDivisiLine'),{{type:'line',data:{{labels:BULAN,datasets:[{{label:'Young Harmonis',data:YH_OMZET,borderColor:'#034543',backgroundColor:'rgba(3,69,67,.08)',fill:true,tension:.4,pointRadius:4}},{{label:'Custom Orders',data:CUST_OMZET,borderColor:'#c2a000',backgroundColor:'rgba(194,160,0,.07)',fill:true,tension:.4,pointRadius:4}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{tooltip:{{callbacks:{{label:c=>' '+fmtM(c.raw)}}}}}},scales:{{y:{{ticks:{{callback:v=>fmtM(v)}},grid:{{color:'#f1f5f9'}}}},x:{{grid:{{display:false}}}}}}}}}});

let cTopOmzet = new Chart(document.getElementById('cTopOmzet'),{{type:'bar',data:{{labels:[],datasets:[{{label:'Omzet',data:[],backgroundColor:'#034543',borderRadius:4}}]}},options:{{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>' '+fmtM(c.raw)}}}}}},scales:{{x:{{ticks:{{callback:v=>fmtM(v)}},grid:{{color:'#f1f5f9'}}}},y:{{grid:{{display:false}},ticks:{{font:{{size:10}}}}}}}}}}}});
let cTopMargin = new Chart(document.getElementById('cTopMargin'),{{type:'bar',data:{{labels:[],datasets:[{{label:'Margin%',data:[],backgroundColor:'#1a8b88',borderRadius:4}}]}},options:{{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>' '+c.raw.toFixed(1)+'%'}}}}}},scales:{{x:{{ticks:{{callback:v=>v+'%'}},max:110,grid:{{color:'#f1f5f9'}}}},y:{{grid:{{display:false}},ticks:{{font:{{size:10}}}}}}}}}}}});

function renderTopProduk(fidx){{
  const prodList = TOP_PRODUK.map(p=>{{
    const omzet  = fidx.reduce((s,i)=>s+(p.monthly_omzet[i]||0),0);
    const qty    = fidx.reduce((s,i)=>s+(p.monthly_qty[i]||0),0);
    const profit = fidx.reduce((s,i)=>s+(p.monthly_profit[i]||0),0);
    const margin = omzet>0?profit/omzet*100:0;
    return {{...p, omzet, qty, profit, margin}};
  }}).filter(p=>p.omzet>0).sort((a,b)=>b.omzet-a.omzet);
  const top10 = prodList.slice(0,10);
  cTopOmzet.data.labels = top10.map(p=>p.produk.length>20?p.produk.slice(0,20)+'…':p.produk);
  cTopOmzet.data.datasets[0].data = top10.map(p=>p.omzet);
  cTopOmzet.update();
  const topMar = [...prodList].sort((a,b)=>b.margin-a.margin).slice(0,10);
  cTopMargin.data.labels = topMar.map(p=>p.produk.length>20?p.produk.slice(0,20)+'…':p.produk);
  cTopMargin.data.datasets[0].data = topMar.map(p=>p.margin);
  cTopMargin.update();
  const catFilter = document.getElementById('topKatFilter')?.value||'all';
  let show = prodList;
  if(catFilter!=='all') show = prodList.filter(p=>p.kategori===catFilter);
  const tb = document.getElementById('tbTop');
  if(tb) tb.innerHTML = show.slice(0,50).map((p,i)=>`<tr><td class="tc">${{i+1}}</td><td><strong>${{p.produk}}</strong></td><td>${{p.varian}}</td><td><span class="badge-blue" style="font-size:.68rem">${{p.kategori.replace('PERLENGKAPAN ','')}}</span></td><td class="tr">${{fmtRp(p.omzet)}}</td><td class="tr" style="font-weight:600">${{p.qty.toLocaleString('id-ID')}}</td><td class="tr">${{fmtRp(p.profit)}}</td><td>${{mBar(p.margin)}}</td></tr>`).join('');
}}

function renderRingkasan(fKatOmzet, fKatProfit, fKatHpp){{
  const tb = document.getElementById('tbKat');
  if(!tb) return;
  const rows = KAT_NAMES.map((nm,i)=>{{
    const omzet=fKatOmzet[i], profit=fKatProfit[i], hpp=fKatHpp[i];
    const margin=omzet>0?profit/omzet*100:0;
    const markup=hpp>0?profit/hpp*100:null;
    return {{no:i+1,nama:nm,omzet,hpp:hpp||null,profit,margin,markup,divisi:KAT_DIVISI[i]}};
  }}).filter(r=>r.omzet>0).sort((a,b)=>b.omzet-a.omzet).map((r,i)=>{{r.no=i+1;return r;}});
  tb.innerHTML = rows.map(r=>`<tr><td class="tc">${{r.no}}</td><td><strong>${{r.nama}}</strong></td><td><span class="badge-blue">${{r.divisi||'—'}}</span></td><td class="tr">${{fmtRp(r.omzet)}}</td><td class="tr">${{r.hpp?fmtRp(r.hpp):'—'}}</td><td class="tr">${{fmtRp(r.profit)}}</td><td>${{mBar(r.margin)}}</td><td class="tr">${{r.markup?pct(r.markup):'—'}}</td></tr>`).join('');
}}

// Initial render
renderTopProduk(FILTERED_IDX);
renderRingkasan(KAT_MONTHLY.map(m=>m.reduce((s,v)=>s+v,0)), KAT_MONTHLY_PROFIT.map(m=>m.reduce((s,v)=>s+v,0)), KAT_MONTHLY_HPP.map(m=>m.reduce((s,v)=>s+v,0)));
let cKatLine=new Chart(document.getElementById('cKatLine'),{{type:'line',data:{{labels:BULAN,datasets:[{{label:KAT_NAMES[0],data:KAT_MONTHLY[0],borderColor:COLORS[0],backgroundColor:'rgba(3,69,67,.08)',fill:true,tension:.4,pointRadius:4}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{tooltip:{{callbacks:{{label:c=>' '+fmtM(c.raw)}}}}}},scales:{{y:{{ticks:{{callback:v=>fmtM(v)}},grid:{{color:'#f1f5f9'}}}},x:{{grid:{{display:false}}}}}}}}}});

function updateKatChart(){{
  const idx=parseInt(document.getElementById('katFilter').value);
  const fBulan = FILTERED_IDX.map(i=>BULAN[i]);
  if(idx===99){{
    cKatLine.data.labels=fBulan;
    cKatLine.data.datasets=KAT_NAMES.map((nm,i)=>({{label:nm.replace('PERLENGKAPAN ','').replace('SERAGAM ',''),data:FILTERED_IDX.map(fi=>KAT_MONTHLY[i][fi]),borderColor:COLORS[i%COLORS.length],backgroundColor:'transparent',tension:.4,pointRadius:2}}));
  }}else{{
    cKatLine.data.labels=fBulan;
    cKatLine.data.datasets=[{{label:KAT_NAMES[idx],data:FILTERED_IDX.map(fi=>KAT_MONTHLY[idx][fi]),borderColor:COLORS[idx%COLORS.length],backgroundColor:COLORS[idx%COLORS.length]+'18',fill:true,tension:.4,pointRadius:4}}];
  }}
  cKatLine.update();
}}

const tbDiv=document.getElementById('tbDiv');
DIVISI_DATA.forEach(d=>{{
  tbDiv.innerHTML+=`<tr><td><strong>${{d.bulan}}</strong></td><td class="tr">${{fmtRp(d.yh_omzet)}}</td><td class="tr">${{fmtRp(d.yh_profit)}}</td><td>${{mBadge(d.yh_margin)}}</td><td class="tr">${{fmtRp(d.cust_omzet)}}</td><td class="tr">${{fmtRp(d.cust_omzet)}}</td><td class="tr">${{pct(d.kontr_yh)}}</td><td class="tr">${{pct(100-d.kontr_yh)}}</td></tr>`;
}});

function filterTopProduk(){{
  renderTopProduk(FILTERED_IDX);
}}

function renderMas(data){{
  const tb=document.getElementById('tbMas');
  tb.innerHTML='';
  data.forEach(m=>{{
    const sc=m.status.includes('Kritis')?'badge-red':m.status.includes('Perhatian')?'badge-orange':'badge-yellow';
    tb.innerHTML+=`<tr><td class="tc">${{m.no}}</td><td><code style="font-size:.78rem;background:#f1f5f9;padding:1px 4px;border-radius:3px">${{m.sku}}</code></td><td><span class="badge-blue" style="font-size:.68rem">${{m.kat.replace('PERLENGKAPAN ','')}}</span></td><td>${{m.produk}}</td><td>${{m.varian}}</td><td>${{m.bulan}}</td><td class="tr">${{m.qty}}</td><td class="tr">${{fmtRp(m.omzet)}}</td><td class="tr">${{fmtRp(m.hpp)}}</td><td class="tr" style="color:#dc2626;font-weight:600">-${{fmtRp(Math.abs(m.profit))}}</td><td class="tr" style="color:#dc2626">${{m.margin.toFixed(1)}}%</td><td><span class="${{sc}}">${{m.status}}</span></td><td style="font-size:.78rem;color:#64748b">${{m.rek}}</td></tr>`;
  }});
}}
renderMas(MASALAH_SKU);
function filterMasalah(){{
  const v=document.getElementById('masFilter').value;
  renderMas(v?MASALAH_SKU.filter(m=>m.kat===v):MASALAH_SKU);
}}

function trendBadge(t,pct){{
  if(t==='naik')return`<span class="badge-green">📈 Naik</span>`;
  if(t==='turun')return`<span class="badge-red">📉 Turun</span>`;
  return`<span class="badge-yellow">➡️ Stabil</span>`;
}}
function rekBadge(qty){{
  return`<span style="background:#fef9c3;color:#854d0e;padding:3px 10px;border-radius:6px;font-weight:700;font-size:.82rem">${{qty}} pcs</span>`;
}}

// ── Stok fisik (toko & gudang) ───────────────────────────────────────────────
let STOK_TOKO   = JSON.parse(localStorage.getItem('stok_toko')||'{{}}');
let STOK_GUDANG = JSON.parse(localStorage.getItem('stok_gudang')||'{{}}');

function uploadStokFisik(files, jenis){{
  if(!files||!files.length) return;
  const doUpload=()=>{{
    const map={{}};
    let loaded=0;
    const total=files.length;
    const tryFinish=()=>{{
      loaded++;
      if(loaded<total) return;
      const n=Object.keys(map).length;
      const lbl='✓ '+(jenis==='toko'?'Toko':'Gudang')+': '+n+' SKU'+(total>1?' ('+total+' file)':'');
      if(jenis==='toko'){{
        STOK_TOKO=map;
        localStorage.setItem('stok_toko',JSON.stringify(map));
        ['stokTokoLabel','invTokoLabel'].forEach(id=>{{ const el=document.getElementById(id); if(el) el.textContent=lbl; }});
      }} else {{
        STOK_GUDANG=map;
        localStorage.setItem('stok_gudang',JSON.stringify(map));
        ['stokGudangLabel','invGudangLabel'].forEach(id=>{{ const el=document.getElementById(id); if(el) el.textContent=lbl; }});
      }}
      filterStok();
      filterInventori();
    }};
    Array.from(files).forEach(file=>{{
      const reader=new FileReader();
      reader.onload=e=>{{
        const wb=XLSX.read(e.target.result,{{type:'array'}});
        const ws=wb.Sheets[wb.SheetNames[0]];
        const rows=XLSX.utils.sheet_to_json(ws,{{header:1}});
        if(!rows.length){{ tryFinish(); return; }}
        const hdr=rows[0].map(h=>String(h||'').toLowerCase().trim());
        const iSku=hdr.indexOf('sku');
        const iQty=hdr.indexOf('stock_qty');
        if(iSku<0||iQty<0){{ tryFinish(); return; }}
        for(let i=1;i<rows.length;i++){{
          const sku=String(rows[i][iSku]||'').trim();
          const qty=Number(rows[i][iQty])||0;
          if(sku) map[sku]=(map[sku]||0)+qty;
        }}
        tryFinish();
      }};
      reader.readAsArrayBuffer(file);
    }});
  }};
  if(typeof XLSX!=='undefined'){{ doUpload(); return; }}
  const s=document.createElement('script');
  s.src='https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js';
  s.onload=doUpload; document.head.appendChild(s);
}}

function updateBudgetRow(input, sku, hpp, hju, perlBeli){{
  const val=Number(input.value)||0;
  const qty=val>0?val:(perlBeli||0);
  const bv=qty*hpp, jv=qty*hju, uv=jv-bv;
  const sid=sku.replace(/[^a-zA-Z0-9]/g,'_');
  const bel=document.getElementById('budget_'+sid);
  const jel=document.getElementById('jual_'+sid);
  const uel=document.getElementById('untung_'+sid);
  if(bel){{ bel.innerHTML=bv>0?fmtRp(bv):(hpp===0?'<span style="color:#94a3b8;font-size:.72rem">HPP?</span>':'—'); bel.dataset.raw=String(bv); }}
  if(jel){{ jel.innerHTML=jv>0?fmtRp(jv):(hju===0?'<span style="color:#94a3b8;font-size:.72rem">Harga?</span>':'—'); jel.dataset.raw=String(jv); }}
  if(uel){{ uel.textContent=uv>0?fmtRp(uv):'—'; uel.dataset.raw=String(uv); }}
  updateTotalBudget();
}}

function updateTotalBudget(){{
  let totalModal=0, totalJual=0;
  document.querySelectorAll('[id^="budget_"]').forEach(el=>{{ totalModal+=Number(el.dataset.raw)||0; }});
  document.querySelectorAll('[id^="jual_"]').forEach(el=>{{ totalJual+=Number(el.dataset.raw)||0; }});
  const profit=totalJual-totalModal;
  const margin=totalJual>0?(profit/totalJual*100):0;
  const bEl=document.getElementById('totalBudgetVal');
  const jEl=document.getElementById('totalJualVal');
  const pEl=document.getElementById('totalProfitVal');
  const mEl=document.getElementById('totalMarginVal');
  if(bEl) bEl.textContent=totalModal>0?fmtRp(totalModal):'—';
  if(jEl) jEl.textContent=totalJual>0?fmtRp(totalJual):'—';
  if(pEl) pEl.textContent=profit>0?fmtRp(profit):'—';
  if(mEl) mEl.textContent=totalJual>0?margin.toFixed(1)+'%':'—';
}}

function stokKey(sku){{ return 'final_order_'+sku; }}
function saveFinal(sku,val){{ localStorage.setItem(stokKey(sku), val); }}
function getFinal(sku){{ return localStorage.getItem(stokKey(sku))||''; }}

// Restore stok label on load
(function(){{
  const nt=Object.keys(STOK_TOKO).length, ng=Object.keys(STOK_GUDANG).length;
  if(nt>0) setTimeout(()=>{{ ['stokTokoLabel','invTokoLabel'].forEach(id=>{{ const el=document.getElementById(id); if(el) el.textContent='✓ Toko: '+nt+' SKU'; }}); }},100);
  if(ng>0) setTimeout(()=>{{ ['stokGudangLabel','invGudangLabel'].forEach(id=>{{ const el=document.getElementById(id); if(el) el.textContent='✓ Gudang: '+ng+' SKU'; }}); }},100);
}})();

function renderStok(data){{
  const tb=document.getElementById('tbStok');
  const cnt=document.getElementById('stokCount');
  tb.innerHTML='';
  cnt.textContent=data.length+' produk ditampilkan';
  data.forEach((s,i)=>{{
    const tpct=s.trend_pct>0?'+'+s.trend_pct.toFixed(0)+'%':s.trend_pct.toFixed(0)+'%';
    const tcolor=s.trend==='naik'?'#15803d':s.trend==='turun'?'#dc2626':'#854d0e';
    const savedFinal=getFinal(s.sku);
    const skuId='final_'+s.sku.replace(/[^a-zA-Z0-9]/g,'_');
    const stToko=STOK_TOKO[s.sku]||0;
    const stGudang=STOK_GUDANG[s.sku]||0;
    const stTotal=stToko+stGudang;
    const perlBeli=Math.max(0,s.rek_qty-stTotal);
    const hpp=s.hpp_unit||0;
    const harga=s.harga_unit||0;
    const finalVal=savedFinal?Number(savedFinal):0;
    const qty4calc=finalVal>0?finalVal:perlBeli;
    const estBudget=qty4calc*hpp;
    const estJual=qty4calc*harga;
    const stOk=stTotal>=s.rek_qty;
    const rowBg=i%2===0?'#fff':'#f8fafc';
    tb.innerHTML+=`<tr>
      <td class="tc" style="position:sticky;left:0;z-index:2;background:${{rowBg}};min-width:38px">${{i+1}}</td>
      <td style="position:sticky;left:38px;z-index:2;background:${{rowBg}};min-width:90px"><code style="font-size:.75rem;background:#f1f5f9;padding:1px 4px;border-radius:3px">${{s.sku}}</code></td>
      <td style="position:sticky;left:128px;z-index:2;background:${{rowBg}};min-width:160px;font-weight:600"><strong>${{s.produk}}</strong></td>
      <td style="position:sticky;left:288px;z-index:2;background:${{rowBg}};min-width:100px;border-right:2px solid #e2e8f0">${{s.varian}}</td>
      <td class="sc-skat"><span class="badge-blue" style="font-size:.68rem">${{s.kategori.replace('PERLENGKAPAN ','')}}</span></td>
      <td class="tr num sc-savg">${{s.avg_per_bulan}}</td>
      <td class="tr num sc-savg3">${{s.recent3_avg}}</td>
      <td class="sc-strnd">${{trendBadge(s.trend)}}</td>
      <td class="tr num sc-strndp" style="color:${{tcolor}}">${{tpct}}</td>
      <td class="tc sc-stipe">${{s.is_seasonal?'<span style="background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe;border-radius:12px;padding:2px 8px;font-size:.72rem;font-weight:700;white-space:nowrap">🗓 Musiman<br><span style="font-weight:400;font-size:.68rem">Puncak: '+s.peak_month+'</span></span>':'<span style="font-size:.78rem;color:#64748b">Reguler</span>'}}</td>
      <td class="tr num sc-sqty">${{s.total_qty.toLocaleString('id-ID')}}</td>
      <td class="tc sc-srek">${{rekBadge(s.rek_qty)}}</td>
      <td class="tr num sc-stoko" style="color:#0369a1">${{stToko>0?stToko:'—'}}</td>
      <td class="tr num sc-sgdng" style="color:#7c3aed">${{stGudang>0?stGudang:'—'}}</td>
      <td class="tr num sc-stotal" style="font-weight:700;color:${{stOk?'#166534':'#dc2626'}}">${{stTotal>0?stTotal:'—'}}</td>
      <td class="tr num sc-sbeli" style="font-weight:700;color:#dc2626">${{perlBeli>0?perlBeli:(stTotal>0?'✓':'—')}}</td>
      <td class="tc sc-slast" style="font-size:.78rem;color:#64748b">${{s.last_sold||'—'}}</td>
      <td class="tc sc-sfinal"><input type="number" min="0" id="${{skuId}}" value="${{savedFinal}}"
        oninput="saveFinal('${{s.sku}}',this.value);updateBudgetRow(this,'${{s.sku}}',${{hpp}},${{harga}},${{perlBeli}})"
        style="width:76px;padding:4px 8px;border:1.5px solid #86efac;border-radius:6px;text-align:right;font-family:'Plus Jakarta Sans',sans-serif;font-size:.85rem;font-weight:600;color:#166534;background:#f0fdf4;outline:none"
        placeholder="0"></td>
      <td class="tr num sc-smodal" id="budget_${{s.sku.replace(/[^a-zA-Z0-9]/g,'_')}}" data-raw="${{estBudget}}" style="color:#854d0e;font-weight:700">${{estBudget>0?fmtRp(estBudget):hpp===0?'<span style=\\'color:#94a3b8;font-size:.72rem\\'>HPP?</span>':'—'}}</td>
      <td class="tr num sc-sjual" id="jual_${{s.sku.replace(/[^a-zA-Z0-9]/g,'_')}}" data-raw="${{estJual}}" style="color:#166534;font-weight:700">${{estJual>0?fmtRp(estJual):harga===0?'<span style=\\'color:#94a3b8;font-size:.72rem\\'>Harga?</span>':'—'}}</td>
      <td class="tr num sc-suntung" id="untung_${{s.sku.replace(/[^a-zA-Z0-9]/g,'_')}}" data-raw="${{estJual-estBudget}}" style="color:#1d4ed8;font-weight:700">${{(estJual-estBudget)>0?fmtRp(estJual-estBudget):'—'}}</td>
    </tr>`;
  }});
}}
renderStok(STOK_DATA);
updateTotalBudget();

function filterStok(){{
  const q=document.getElementById('stokSearch').value.toLowerCase();
  const kat=document.getElementById('stokKatFilter').value;
  const trend=document.getElementById('stokTrendFilter').value;
  const beli=document.getElementById('stokBeliFilter').value;
  let d=STOK_DATA.filter(s=>{{
    if(q&&!s.produk.toLowerCase().includes(q)&&!s.sku.toLowerCase().includes(q)&&!s.varian.toLowerCase().includes(q))return false;
    if(kat&&s.kategori!==kat)return false;
    if(trend==='musiman'){{if(!s.is_seasonal)return false;}}
    else if(trend&&s.trend!==trend)return false;
    if(beli==='beli'){{
      const tot=(STOK_TOKO[s.sku]||0)+(STOK_GUDANG[s.sku]||0);
      if(tot>=s.rek_qty)return false;
    }}
    if(beli==='cukup'){{
      const tot=(STOK_TOKO[s.sku]||0)+(STOK_GUDANG[s.sku]||0);
      if(tot<s.rek_qty)return false;
    }}
    return true;
  }});
  renderStok(d);
  updateTotalBudget();
}}

function toggleStokKolomPanel(e){{
  e.stopPropagation();
  const p=document.getElementById('stokKolomPanel');
  p.style.display=p.style.display==='block'?'none':'block';
}}
document.addEventListener('click',function(e){{
  const p=document.getElementById('stokKolomPanel');
  if(p&&!p.contains(e.target)&&e.target.id!=='btnStokKolom')p.style.display='none';
}});
function toggleStokCol(key,visible){{
  const wrap=document.getElementById('stokTableWrap');
  if(!wrap)return;
  if(visible)wrap.classList.remove('stok-hide-'+key);
  else wrap.classList.add('stok-hide-'+key);
}}
function stokKolomAll(visible){{
  document.querySelectorAll('#stokKolomPanel input[type=checkbox]').forEach(function(cb){{
    cb.checked=visible;
    if(cb.dataset.col)toggleStokCol(cb.dataset.col,visible);
  }});
}}

function stokTableData(){{
  const headers=['No','SKU','Produk','Varian','Kategori','Total Qty','Avg/Bln','Avg 3Bln Terakhir','Tren','Tren %','Tipe','Puncak','Rek. Stok','Stok Toko','Stok Gudang','Total Stok','Perlu Beli','Terakhir Terjual','Final Order','Est. Modal','Est. Nilai Jual','Est. Untung'];
  const rows=STOK_DATA.map((s,i)=>{{
    const tpct=(s.trend_pct>0?'+':'')+s.trend_pct.toFixed(0)+'%';
    const final=getFinal(s.sku);
    const finalNum=final?Number(final):0;
    const tipe=s.is_seasonal?'Musiman':'Reguler';
    const stToko=STOK_TOKO[s.sku]||0;
    const stGudang=STOK_GUDANG[s.sku]||0;
    const stTotal=stToko+stGudang;
    const perlBeli=Math.max(0,s.rek_qty-stTotal);
    const hpp=s.hpp_unit||0;
    const harga=s.harga_unit||0;
    const qty4c=finalNum>0?finalNum:perlBeli;
    const modal=qty4c*hpp, jual=qty4c*harga, untung=jual-modal;
    return [i+1,s.sku,s.produk,s.varian,s.kategori,s.total_qty,s.avg_per_bulan,s.recent3_avg,s.trend,tpct,tipe,s.peak_month||'',s.rek_qty,stToko||'',stGudang||'',stTotal||'',perlBeli||'',s.last_sold||'',finalNum||'',modal||'',jual||'',untung||''];
  }});
  return {{headers,rows}};
}}

function downloadStokCSV(){{
  const {{headers,rows}}=stokTableData();
  const esc=v=>{{const s=String(v);return s.includes(',')||s.includes('"')?'"'+s.replace(/"/g,'""')+'"':s;}};
  const nl='\\r\\n';
  const csv=[headers.map(esc).join(','),...rows.map(r=>r.map(esc).join(','))].join(nl);
  const blob=new Blob(['﻿'+csv],{{type:'text/csv;charset=utf-8'}});
  const a=document.createElement('a');a.href=URL.createObjectURL(blob);
  a.download='Rekomendasi_Stok_Harmoni.csv';a.click();
}}

function downloadStokXLSX(){{
  const doExport=()=>{{
    const {{headers,rows}}=stokTableData();
    const wb=XLSX.utils.book_new();
    const ws=XLSX.utils.aoa_to_sheet([headers,...rows]);
    // Column widths
    ws['!cols']=[{{wch:4}},{{wch:18}},{{wch:30}},{{wch:20}},{{wch:22}},{{wch:12}},{{wch:10}},{{wch:16}},{{wch:8}},{{wch:8}},{{wch:10}},{{wch:16}},{{wch:12}}];
    XLSX.utils.book_append_sheet(wb,ws,'Rekomendasi Stok');
    XLSX.writeFile(wb,'Rekomendasi_Stok_Harmoni.xlsx');
  }};
  if(typeof XLSX!=='undefined'){{ doExport(); return; }}
  const s=document.createElement('script');
  s.src='https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js';
  s.onload=doExport;
  document.head.appendChild(s);
}}
</script>
</body>
</html>"""

with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"✅ HTML berhasil dibuat: {OUTPUT_PATH}")
print(f"   Ukuran: {os.path.getsize(OUTPUT_PATH):,} bytes")
print(f"\n👉 Buka file ini di browser: {OUTPUT_PATH}")
